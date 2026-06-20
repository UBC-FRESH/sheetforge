import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from sheetforge.extraction import extract_workbook
from sheetforge.graph import DependencyGraph, build_dependency_graph
from tests.fixtures.synthetic_model.build_workbook import build_workbook


def test_build_dependency_graph_recovers_synthetic_execution_edges(tmp_path: Path) -> None:
    workbook = extract_workbook(build_workbook(tmp_path / "synthetic_model.xlsx"))

    graph = build_dependency_graph(workbook)

    assert [(edge.source.normalized, edge.target.normalized) for edge in graph.execution_edges] == [
        ("Inputs!B2", "Calc!B2"),
        ("Inputs!B3", "Calc!B2"),
        ("Calc!B2", "Calc!B3"),
        ("Inputs!B4", "Calc!B3"),
        ("Calc!B3", "Calc!B4"),
        ("Calc!B4", "Summary!B2"),
        ("Summary!B2", "Summary!B3"),
    ]


def test_build_dependency_graph_preserves_semantic_named_range_edges(tmp_path: Path) -> None:
    workbook = extract_workbook(build_workbook(tmp_path / "synthetic_model.xlsx"))

    graph = build_dependency_graph(workbook)
    semantic_edges = [
        edge for edge in graph.semantic_edges if edge.target.normalized == "Calc!B2"
    ]
    execution_edges = [
        edge for edge in graph.execution_edges if edge.target.normalized == "Calc!B2"
    ]

    assert [(edge.source.kind, edge.source.normalized, edge.raw_reference) for edge in semantic_edges] == [
        ("named_range", "BaseVolume", "BaseVolume"),
        ("named_range", "GrowthRate", "GrowthRate"),
    ]
    assert [(edge.source.normalized, edge.resolved_from.normalized) for edge in execution_edges if edge.resolved_from] == [
        ("Inputs!B2", "BaseVolume"),
        ("Inputs!B3", "GrowthRate"),
    ]


def test_dependency_graph_payload_round_trips(tmp_path: Path) -> None:
    workbook = extract_workbook(build_workbook(tmp_path / "synthetic_model.xlsx"))

    graph = build_dependency_graph(workbook)
    payload = graph.to_dict()

    assert json.loads(json.dumps(payload)) == payload
    assert DependencyGraph.from_dict(payload) == graph


def test_dependency_graph_reports_external_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "external.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["B2"] = "='[other.xlsx]Inputs'!B2"
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))

    assert graph.diagnostics == ("external_reference",)
    assert graph.execution_edges[0].diagnostic_code == "external_reference"
    assert graph.execution_edges[0].source.kind == "external"


def test_dependency_graph_reports_structured_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["A1"] = "Amount"
    sheet["A2"] = 10
    sheet["B1"] = "=Table1[Amount]"
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))

    assert graph.diagnostics == ("unsupported_structured_reference",)
    assert graph.semantic_edges[0].source.kind == "structured"
    assert graph.semantic_edges[0].diagnostic_code == "unsupported_structured_reference"
    assert graph.execution_edges[0].source.kind == "structured"
    assert graph.execution_edges[0].diagnostic_code == "unsupported_structured_reference"


def test_dependency_graph_resolves_current_row_structured_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "current-row-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, "=InputTable[[#This Row],[Amount]]"])
    sheet.append([20, "=InputTable[[#This Row],[Amount]]"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))
    execution_edges = [edge for edge in graph.execution_edges if edge.target.normalized == "Data!B2"]

    assert graph.diagnostics == ()
    assert [(edge.source.normalized, edge.resolved_from.normalized) for edge in execution_edges if edge.resolved_from] == [
        ("Data!A2", "InputTable[[#This Row],[Amount]]")
    ]


def test_dependency_graph_resolves_column_structured_reference_as_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "column-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Rate"])
    sheet.append([10, 0.1])
    sheet.append([20, 0.2])
    sheet["D1"] = "=SUM(InputTable[Amount])"
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))
    execution_edges = [edge for edge in graph.execution_edges if edge.target.normalized == "Data!D1"]

    assert graph.diagnostics == ()
    assert [(edge.source.kind, edge.source.normalized, edge.resolved_from.normalized) for edge in execution_edges if edge.resolved_from] == [
        ("range", "Data!A2:A3", "InputTable[Amount]")
    ]


def test_dependency_graph_resolves_whole_table_structured_references_as_ranges(tmp_path: Path) -> None:
    workbook_path = tmp_path / "whole-table-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Rate"])
    sheet.append([10, 0.1])
    sheet.append([20, 0.2])
    sheet["D1"] = "=SUM(InputTable[])"
    sheet["D2"] = "=SUM(InputTable[#All])"
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))

    data_body_edges = [edge for edge in graph.execution_edges if edge.target.normalized == "Data!D1"]
    all_edges = [edge for edge in graph.execution_edges if edge.target.normalized == "Data!D2"]
    assert graph.diagnostics == ()
    assert [(edge.source.kind, edge.source.normalized, edge.resolved_from.normalized) for edge in data_body_edges if edge.resolved_from] == [
        ("range", "Data!A2:B3", "InputTable[]")
    ]
    assert [(edge.source.kind, edge.source.normalized, edge.resolved_from.normalized) for edge in all_edges if edge.resolved_from] == [
        ("range", "Data!A1:B3", "InputTable[#All]")
    ]


def test_dependency_graph_resolves_cross_table_current_row_by_row_offset(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cross-table-current-row.xlsx"
    source = Workbook()
    source_sheet = source.active
    source_sheet.title = "Source"
    source_sheet.append(["Amount"])
    source_sheet.append([10])
    source_sheet.append([20])
    source_sheet.add_table(Table(displayName="SourceTable", ref="A1:A3"))
    target_sheet = source.create_sheet("Target")
    target_sheet.append(["Result"])
    target_sheet.append(["=SourceTable[[#This Row],[Amount]]"])
    target_sheet.append(["=SourceTable[[#This Row],[Amount]]"])
    target_sheet.add_table(Table(displayName="TargetTable", ref="A1:A3"))
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))

    execution_edges = [edge for edge in graph.execution_edges if edge.target.normalized in {"Target!A2", "Target!A3"}]
    assert graph.diagnostics == ()
    assert [(edge.source.normalized, edge.target.normalized, edge.resolved_from.normalized) for edge in execution_edges if edge.resolved_from] == [
        ("Source!A2", "Target!A2", "SourceTable[[#This Row],[Amount]]"),
        ("Source!A3", "Target!A3", "SourceTable[[#This Row],[Amount]]"),
    ]


def test_dependency_graph_expands_range_execution_edges(tmp_path: Path) -> None:
    workbook_path = tmp_path / "range.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Inputs"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["B1"] = "=SUM(A1:A2)"
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))

    range_execution_edges = [edge for edge in graph.execution_edges if edge.resolved_from is not None]
    assert [(edge.source.normalized, edge.target.normalized) for edge in range_execution_edges] == [
        ("Inputs!A1", "Inputs!B1"),
        ("Inputs!A2", "Inputs!B1"),
    ]
    assert {edge.resolved_from.normalized for edge in range_execution_edges if edge.resolved_from} == {
        "Inputs!A1:A2"
    }


def test_dependency_graph_reports_simple_circular_dependency(tmp_path: Path) -> None:
    workbook_path = tmp_path / "circular.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "=A2"
    sheet["A2"] = "=A1"
    source.save(workbook_path)

    graph = build_dependency_graph(extract_workbook(workbook_path))

    assert "circular_dependency" in graph.diagnostics
