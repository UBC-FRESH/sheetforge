from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from sheetforge.extraction import CellRecord, extract_workbook
from sheetforge.formulas import build_formula_reference_index, translate_formula_cell
from sheetforge.graph import build_dependency_graph
from tests.fixtures.synthetic_model.build_workbook import build_workbook


def synthetic_formula_cells(tmp_path: Path) -> tuple[dict[str, CellRecord], object]:
    workbook = extract_workbook(build_workbook(tmp_path / "synthetic_model.xlsx"))
    graph = build_dependency_graph(workbook)
    return {cell.cell_ref: cell for cell in workbook.cells if cell.formula is not None}, graph


def test_translate_named_range_arithmetic_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Calc!B2"], graph)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "binary"
    assert expression.root.operator == "*"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Inputs!B2"
    assert expression.root.operands[1].kind == "binary"
    assert expression.root.operands[1].operator == "+"
    assert expression.root.operands[1].operands[0].value == 1
    assert expression.root.operands[1].operands[1].reference is not None
    assert expression.root.operands[1].operands[1].reference.normalized == "Inputs!B3"


def test_translate_formula_uses_reference_index(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)
    reference_index = build_formula_reference_index(graph)

    expression = translate_formula_cell(cells["Calc!B2"], graph, reference_index=reference_index)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Inputs!B2"


def test_translate_sheet_relative_arithmetic_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Calc!B3"], graph)

    assert expression.root is not None
    assert expression.root.operator == "*"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Calc!B2"
    assert expression.root.operands[1].reference is not None
    assert expression.root.operands[1].reference.normalized == "Inputs!B4"


def test_translate_round_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Calc!B4"], graph)

    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "ROUND"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Calc!B3"
    assert expression.root.operands[1].value == 2


def test_translate_direct_reference_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Summary!B2"], graph)

    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Calc!B4"


def test_translate_if_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Summary!B3"], graph)

    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "IF"
    assert expression.root.operands[0].kind == "comparison"
    assert expression.root.operands[0].operator == ">"
    assert expression.root.operands[0].operands[0].reference is not None
    assert expression.root.operands[0].operands[0].reference.normalized == "Summary!B2"
    assert expression.root.operands[0].operands[1].value == 50
    assert expression.root.operands[1].value == "ok"
    assert expression.root.operands[2].value == "low"


def test_translate_unsupported_function_reports_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "unsupported_function.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "needle"
    sheet["B1"] = "=XLOOKUP(A1,A:A,B:B)"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_function"
    assert expression.diagnostics[0].severity == "error"
    assert expression.diagnostics[0].raw_value == "XLOOKUP"


def test_translate_unsupported_operator_reports_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "unsupported_operator.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 2
    sheet["B1"] = "=A1%"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_operator"
    assert expression.diagnostics[0].severity == "error"
    assert expression.diagnostics[0].raw_value == "%"


def test_translate_structured_reference_reports_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["A1"] = "Amount"
    sheet["A2"] = 10
    sheet["B1"] = "=Table1[Amount]"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_structured_reference"
    assert expression.diagnostics[0].severity == "error"
    assert expression.diagnostics[0].raw_value == "Table1[Amount]"


def test_translate_current_row_structured_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "current-row-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, "=InputTable[[#This Row],[Amount]]"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B2"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B2")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Data!A2"


def test_translate_column_structured_reference_as_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "column-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Rate"])
    sheet.append([10, 0.1])
    sheet.append([20, 0.2])
    sheet["D1"] = "=SUM(InputTable[Amount])"
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!D1")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "SUM"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.kind == "range"
    assert expression.root.operands[0].reference.normalized == "Data!A2:A3"


def test_translate_boolean_literal(tmp_path: Path) -> None:
    workbook_path = tmp_path / "boolean-literal.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "=FALSE"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!A1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "literal"
    assert expression.root.value is False


def test_translate_unary_minus_exponent_and_concat(tmp_path: Path) -> None:
    workbook_path = tmp_path / "operators.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 3
    sheet["A2"] = "x"
    sheet["B1"] = "=-A1"
    sheet["B2"] = "=A1^2"
    sheet["B3"] = '=A2&"y"'
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cells = {cell.cell_ref: cell for cell in workbook.cells if cell.formula is not None}

    unary = translate_formula_cell(formula_cells["Calc!B1"], graph)
    exponent = translate_formula_cell(formula_cells["Calc!B2"], graph)
    concat = translate_formula_cell(formula_cells["Calc!B3"], graph)

    assert unary.translated is True
    assert unary.root is not None
    assert unary.root.kind == "unary"
    assert unary.root.operator == "-"
    assert exponent.translated is True
    assert exponent.root is not None
    assert exponent.root.kind == "binary"
    assert exponent.root.operator == "^"
    assert concat.translated is True
    assert concat.root is not None
    assert concat.root.kind == "binary"
    assert concat.root.operator == "&"


def test_translate_ref_error_reports_sharp_diagnostic(tmp_path: Path) -> None:
    workbook_path = tmp_path / "ref-error.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "=#REF!"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!A1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_error_reference"
    assert expression.diagnostics[0].raw_value == "#REF!"


def test_translate_static_offset_to_concrete_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "static-offset.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, None])
    sheet.append([20, "=OFFSET(InputTable[[#This Row],[Amount]],-1,0)"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B3")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Data!A2"


def test_translate_dynamic_offset_shape_reports_sharp_diagnostic(tmp_path: Path) -> None:
    workbook_path = tmp_path / "dynamic-offset.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 10
    sheet["B1"] = "=OFFSET(A1,0,0,2,1)"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_offset_shape"
    assert expression.diagnostics[0].raw_value == "OFFSET"
