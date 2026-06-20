import json
from pathlib import Path

from openpyxl import load_workbook

from tests.fixtures.synthetic_model import build_workbook


FIXTURE_ROOT = Path(__file__).parent / "fixtures" / "synthetic_model"


def test_synthetic_workbook_builder_writes_expected_workbook(tmp_path: Path) -> None:
    workbook_path = build_workbook(tmp_path / "synthetic_model.xlsx")

    workbook = load_workbook(workbook_path, data_only=False)

    assert workbook.sheetnames == ["Inputs", "Calc", "Summary"]
    assert workbook["Inputs"]["B2"].value == 100
    assert workbook["Inputs"]["B3"].value == 0.08
    assert workbook["Inputs"]["B4"].value == 0.65
    assert workbook["Calc"]["B2"].value == "=BaseVolume*(1+GrowthRate)"
    assert workbook["Calc"]["B3"].value == "=B2*Inputs!B4"
    assert workbook["Calc"]["B4"].value == "=ROUND(B3,2)"
    assert workbook["Summary"]["B2"].value == "=Calc!B4"
    assert workbook["Summary"]["B3"].value == '=IF(B2>50,"ok","low")'

    defined_names = workbook.defined_names
    assert list(defined_names["BaseVolume"].destinations) == [("Inputs", "$B$2")]
    assert list(defined_names["GrowthRate"].destinations) == [("Inputs", "$B$3")]


def test_synthetic_fixture_json_assets_are_stable() -> None:
    scenario = json.loads((FIXTURE_ROOT / "baseline_scenario.json").read_text(encoding="utf-8"))
    expected = json.loads((FIXTURE_ROOT / "baseline_expected_outputs.json").read_text(encoding="utf-8"))

    assert scenario["scenario_id"] == "synthetic_model_baseline"
    assert scenario["outputs"] == [
        {"cell_ref": "Summary!B2", "kind": "number", "tolerance": 1e-9},
        {"cell_ref": "Summary!B3", "kind": "text"},
    ]
    assert expected["outputs"]["Summary!B2"] == {
        "kind": "number",
        "value": 70.2,
        "tolerance": 1e-9,
    }
    assert expected["outputs"]["Summary!B3"] == {"kind": "text", "value": "ok"}
