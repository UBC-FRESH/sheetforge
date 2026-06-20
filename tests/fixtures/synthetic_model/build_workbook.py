from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import quote_sheetname
from openpyxl.workbook.defined_name import DefinedName


def build_workbook(path: str | Path) -> Path:
    """Write the controlled synthetic workbook to path and return that path."""

    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    calc = workbook.create_sheet("Calc")
    summary = workbook.create_sheet("Summary")

    inputs["A1"] = "Parameter"
    inputs["B1"] = "Value"
    inputs["A2"] = "Base volume"
    inputs["B2"] = 100
    inputs["A3"] = "Growth rate"
    inputs["B3"] = 0.08
    inputs["A4"] = "Harvest share"
    inputs["B4"] = 0.65

    calc["A1"] = "Metric"
    calc["B1"] = "Formula"
    calc["A2"] = "Grown volume"
    calc["B2"] = "=BaseVolume*(1+GrowthRate)"
    calc["A3"] = "Harvest volume"
    calc["B3"] = "=B2*Inputs!B4"
    calc["A4"] = "Rounded harvest"
    calc["B4"] = "=ROUND(B3,2)"

    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Output"
    summary["B2"] = "=Calc!B4"
    summary["A3"] = "Status"
    summary["B3"] = '=IF(B2>50,"ok","low")'

    workbook.defined_names.add(DefinedName("BaseVolume", attr_text=f"{quote_sheetname('Inputs')}!$B$2"))
    workbook.defined_names.add(DefinedName("GrowthRate", attr_text=f"{quote_sheetname('Inputs')}!$B$3"))

    workbook.save(workbook_path)
    return workbook_path
