import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName

from sheetforge.conversion import ConversionPlan, build_conversion_plan
from sheetforge.extraction import extract_workbook
from sheetforge.formulas import build_formula_reference_index, translate_formula_cell
from sheetforge.generation import GeneratedModuleContract, GeneratedSymbol, generate_python_module
from sheetforge.graph import build_dependency_graph
from sheetforge.validation import ComparisonRules, OracleConfig, ScenarioOutput, ValidationScenario, build_validation_report
from tests.fixtures.synthetic_model.build_workbook import build_workbook


def test_build_conversion_plan_summarizes_clean_synthetic_workflow(tmp_path: Path) -> None:
    workbook = extract_workbook(build_workbook(tmp_path / "synthetic_model.xlsx"))
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)
    formula_cells = {cell.cell_ref: cell for cell in workbook.cells if cell.formula is not None}
    expressions = {
        cell_ref: translate_formula_cell(cell, graph, reference_index=reference_index)
        for cell_ref, cell in formula_cells.items()
    }
    contract = GeneratedModuleContract(
        workbook_id=workbook.workbook_id,
        module_name="synthetic_model",
        input_refs=("Inputs!B2", "Inputs!B3", "Inputs!B4"),
        output_refs=("Summary!B2",),
        symbols=(
            GeneratedSymbol(cell_ref="Inputs!B2", symbol_name="inputs_b2", kind="input"),
            GeneratedSymbol(cell_ref="Inputs!B3", symbol_name="inputs_b3", kind="input"),
            GeneratedSymbol(cell_ref="Inputs!B4", symbol_name="inputs_b4", kind="input"),
            GeneratedSymbol(cell_ref="Calc!B2", symbol_name="calc_b2", kind="intermediate", raw_formula="=BaseVolume*(1+GrowthRate)"),
            GeneratedSymbol(cell_ref="Calc!B3", symbol_name="calc_b3", kind="intermediate", raw_formula="=Calc!B2*AdjustmentFactor"),
            GeneratedSymbol(cell_ref="Calc!B4", symbol_name="calc_b4", kind="intermediate", raw_formula="=ROUND(B3,2)"),
            GeneratedSymbol(cell_ref="Summary!B2", symbol_name="summary_b2", kind="output", raw_formula="=Calc!B4"),
        ),
    )
    generation_result = generate_python_module(
        contract=contract,
        expressions=expressions,
        constants={"Inputs!B2": 100, "Inputs!B3": 0.1, "Inputs!B4": 0.9},
    )
    scenario = ValidationScenario(
        scenario_id="synthetic",
        description="synthetic",
        source_workbook=workbook.source_path,
        generated_model="generated.py",
        oracle=OracleConfig(backend="cached_workbook"),
        inputs=(),
        outputs=(ScenarioOutput(cell_ref="Summary!B2", kind="number"),),
        comparison=ComparisonRules(),
    )
    validation_report = build_validation_report(
        scenario=scenario,
        generated_values={"Summary!B2": 99.0},
        oracle_values={"Summary!B2": 99.0},
    )

    plan = build_conversion_plan(
        plan_id="synthetic-plan",
        workbook=workbook,
        graph=graph,
        expressions=expressions,
        benchmark_role="synthetic_fixture",
        generation_result=generation_result,
        cached_validation_report=validation_report,
        oracle_blockers=("oracle_calculation_failed",),
    )
    payload = plan.to_dict()

    assert plan.coverage.formula_cells == 5
    assert plan.coverage.translated_formula_cells == 5
    assert plan.coverage.translation_coverage == 1.0
    assert plan.generation.generated is True
    assert plan.generation.selected_outputs == 1
    assert plan.validation.cached_validation_status == "pass"
    assert plan.workflow_status.overall == "partial"
    blockers_by_code = {blocker.diagnostic_code: blocker for blocker in plan.residual_blockers}
    assert blockers_by_code["missing_cached_formula_value"].category == "missing_cached_values"
    assert blockers_by_code["missing_cached_formula_value"].disposition == "deferred"
    assert blockers_by_code["oracle_calculation_failed"].category == "validation_oracle"
    assert payload["privacy_review"]["contains_source_path"] is False
    assert ConversionPlan.from_dict(json.loads(json.dumps(payload))) == plan


def test_build_conversion_plan_classifies_translation_blockers(tmp_path: Path) -> None:
    workbook_path = tmp_path / "unsupported.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 1
    sheet["B1"] = "=XLOOKUP(A1,A:A,A:A)"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.formula is not None)
    expressions = {formula_cell.cell_ref: translate_formula_cell(formula_cell, graph)}

    plan = build_conversion_plan(
        plan_id="unsupported-plan",
        workbook=workbook,
        graph=graph,
        expressions=expressions,
        benchmark_role="synthetic_fixture",
    )

    assert plan.coverage.formula_cells == 1
    assert plan.coverage.translated_formula_cells == 0
    assert plan.diagnostic_summary.translation == {"unsupported_function": 1}
    blockers_by_code = {blocker.diagnostic_code: blocker for blocker in plan.residual_blockers}
    assert blockers_by_code["unsupported_function"].category == "unsupported_formula_semantics"
    assert blockers_by_code["unsupported_function"].disposition == "next_target"
    assert plan.recommendations[0].action == "Implement support or a sharper diagnostic for this translation blocker."


def test_build_conversion_plan_classifies_named_range_source_errors(tmp_path: Path) -> None:
    workbook_path = tmp_path / "named-range-source-error.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 1
    source.defined_names.add(DefinedName("BrokenName", attr_text="#REF!"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)

    plan = build_conversion_plan(
        plan_id="named-range-source-error-plan",
        workbook=workbook,
        graph=graph,
        expressions={},
        benchmark_role="synthetic_fixture",
    )
    blockers_by_code = {blocker.diagnostic_code: blocker for blocker in plan.residual_blockers}

    assert plan.diagnostic_summary.named_ranges == {"named_range_source_error": 1}
    assert blockers_by_code["named_range_source_error"].category == "source_workbook_defect"
    assert blockers_by_code["named_range_source_error"].disposition == "out_of_scope"


def test_build_conversion_plan_classifies_external_formula_references(tmp_path: Path) -> None:
    workbook_path = tmp_path / "external-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "='[external.xlsx]Inputs'!A1"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.formula is not None)
    expressions = {formula_cell.cell_ref: translate_formula_cell(formula_cell, graph)}

    plan = build_conversion_plan(
        plan_id="external-reference-plan",
        workbook=workbook,
        graph=graph,
        expressions=expressions,
        benchmark_role="synthetic_fixture",
    )
    blockers_by_code = {blocker.diagnostic_code: blocker for blocker in plan.residual_blockers}

    assert blockers_by_code["unsupported_external_link"].category == "external_dependency"
    assert blockers_by_code["unsupported_external_link"].disposition == "deferred"
    assert blockers_by_code["external_reference"].category == "external_dependency"
    assert blockers_by_code["external_reference"].disposition == "deferred"


def test_build_conversion_plan_marks_resolved_structured_reference_provenance(tmp_path: Path) -> None:
    workbook_path = tmp_path / "resolved-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, "=InputTable[[#This Row],[Amount]]"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B2"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)
    formula_cell = next(cell for cell in workbook.cells if cell.formula is not None)
    expressions = {formula_cell.cell_ref: translate_formula_cell(formula_cell, graph, reference_index)}

    plan = build_conversion_plan(
        plan_id="resolved-structured-reference-plan",
        workbook=workbook,
        graph=graph,
        expressions=expressions,
        benchmark_role="synthetic_fixture",
    )
    blockers_by_code = {blocker.diagnostic_code: blocker for blocker in plan.residual_blockers}

    assert plan.diagnostic_summary.formula_extraction == {
        "missing_cached_formula_value": 1,
        "unsupported_structured_reference": 1,
    }
    assert plan.diagnostic_summary.graph == {}
    assert plan.diagnostic_summary.translation == {}
    assert blockers_by_code["unsupported_structured_reference"].disposition == "resolved"
    assert blockers_by_code["unsupported_structured_reference"].next_action.startswith("No conversion action required")


def test_build_conversion_plan_marks_resolved_static_offset_volatility(tmp_path: Path) -> None:
    workbook_path = tmp_path / "resolved-static-offset-volatility.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, None])
    sheet.append([20, "=OFFSET(InputTable[[#This Row],[Amount]],-1,0)"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)
    formula_cell = next(cell for cell in workbook.cells if cell.formula is not None)
    expressions = {formula_cell.cell_ref: translate_formula_cell(formula_cell, graph, reference_index)}

    plan = build_conversion_plan(
        plan_id="resolved-static-offset-volatility-plan",
        workbook=workbook,
        graph=graph,
        expressions=expressions,
        benchmark_role="synthetic_fixture",
    )
    blockers_by_code = {blocker.diagnostic_code: blocker for blocker in plan.residual_blockers}

    assert plan.diagnostic_summary.formula_extraction == {
        "missing_cached_formula_value": 1,
        "unsupported_structured_reference": 1,
        "unsupported_volatile_function": 1,
    }
    assert plan.diagnostic_summary.translation == {}
    assert blockers_by_code["unsupported_volatile_function"].disposition == "resolved"
    assert blockers_by_code["unsupported_volatile_function"].next_action.startswith("No formula-semantics action required")
    assert blockers_by_code["missing_cached_formula_value"].disposition == "deferred"
