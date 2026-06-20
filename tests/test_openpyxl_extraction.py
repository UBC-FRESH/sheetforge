import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName

from sheetforge.extraction import extract_workbook
from tests.fixtures.synthetic_model.build_workbook import build_workbook


def test_extract_workbook_reads_synthetic_fixture(tmp_path: Path) -> None:
    workbook_path = build_workbook(tmp_path / "synthetic_model.xlsx")

    workbook = extract_workbook(workbook_path)

    assert workbook.workbook_id == "synthetic_model.xlsx"
    assert workbook.source_path == str(workbook_path)
    assert [sheet.sheet_id for sheet in workbook.sheets] == ["Inputs", "Calc", "Summary"]
    assert [sheet.state for sheet in workbook.sheets] == ["visible", "visible", "visible"]
    assert len(workbook.cells) == 22
    assert workbook.diagnostics == ()

    named_ranges = {named_range.name: named_range for named_range in workbook.named_ranges}
    assert sorted(named_ranges) == ["BaseVolume", "GrowthRate"]
    assert named_ranges["BaseVolume"].scope == "workbook"
    assert named_ranges["BaseVolume"].raw_definition == "'Inputs'!$B$2"
    assert named_ranges["BaseVolume"].destinations == ("Inputs!B2",)
    assert named_ranges["BaseVolume"].status == "resolved"

    cells = {cell.cell_ref: cell for cell in workbook.cells}
    assert cells["Inputs!B2"].kind == "value"
    assert cells["Inputs!B2"].raw_value == 100
    assert cells["Inputs!B2"].data_type == "n"
    assert cells["Inputs!B2"].formula is None
    assert cells["Calc!B2"].kind == "formula"
    assert cells["Calc!B2"].raw_value == "=BaseVolume*(1+GrowthRate)"
    assert cells["Calc!B2"].formula is not None
    assert cells["Calc!B2"].formula.raw_references == ("BaseVolume", "GrowthRate")
    assert cells["Calc!B2"].formula.functions == ()
    assert cells["Calc!B4"].formula is not None
    assert cells["Calc!B4"].formula.functions == ("ROUND",)
    assert cells["Summary!B3"].formula is not None
    assert cells["Summary!B3"].formula.functions == ("IF",)


def test_extract_workbook_emits_missing_cached_formula_diagnostics(tmp_path: Path) -> None:
    workbook_path = build_workbook(tmp_path / "synthetic_model.xlsx")

    workbook = extract_workbook(workbook_path)
    formula_cells = [cell for cell in workbook.cells if cell.kind == "formula"]
    diagnostics = [
        diagnostic
        for cell in formula_cells
        if cell.formula is not None
        for diagnostic in cell.formula.diagnostics
    ]

    assert [cell.cell_ref for cell in formula_cells] == [
        "Calc!B2",
        "Calc!B3",
        "Calc!B4",
        "Summary!B2",
        "Summary!B3",
    ]
    assert len(diagnostics) == 5
    assert {diagnostic.code for diagnostic in diagnostics} == {"missing_cached_formula_value"}
    assert diagnostics[0].severity == "warning"


def test_extract_workbook_payload_round_trips(tmp_path: Path) -> None:
    workbook_path = build_workbook(tmp_path / "synthetic_model.xlsx")

    workbook = extract_workbook(workbook_path)
    payload = workbook.to_dict()

    assert json.loads(json.dumps(payload)) == payload
    assert type(workbook).from_dict(payload) == workbook


def test_extract_workbook_reports_volatile_formula(tmp_path: Path) -> None:
    workbook_path = tmp_path / "volatile.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Inputs"
    sheet["A1"] = "=NOW()"
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)
    cell = workbook.cells[0]

    assert cell.cell_ref == "Inputs!A1"
    assert cell.formula is not None
    assert {diagnostic.code for diagnostic in cell.formula.diagnostics} == {
        "missing_cached_formula_value",
        "unsupported_volatile_function",
    }


def test_extract_workbook_reports_structured_reference_formula(tmp_path: Path) -> None:
    workbook_path = tmp_path / "structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["A1"] = "Amount"
    sheet["A2"] = 10
    sheet["B1"] = "=Table1[Amount]"
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)
    cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B1")

    assert cell.formula is not None
    assert cell.formula.raw_references == ("Table1[Amount]",)
    assert {diagnostic.code for diagnostic in cell.formula.diagnostics} == {
        "missing_cached_formula_value",
        "unsupported_structured_reference",
    }


def test_extract_workbook_reads_table_metadata(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Rate", "Result"])
    sheet.append([10, 0.1, None])
    sheet.append([20, 0.2, None])
    sheet.add_table(Table(displayName="InputTable", ref="A1:C3"))
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)

    assert [table.to_dict() for table in workbook.tables] == [
        {
            "name": "InputTable",
            "sheet": "Data",
            "ref": "A1:C3",
            "columns": ["Amount", "Rate", "Result"],
        }
    ]


def test_extract_workbook_resolves_structured_defined_name_to_table_column_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "structured-defined-name.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Product", "Amount"])
    sheet.append(["A", 10])
    sheet.append(["B", 20])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.defined_names.add(DefinedName("ProductList", attr_text="InputTable[[#Data],[Product]]"))
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)
    named_range = workbook.named_ranges[0]

    assert named_range.name == "ProductList"
    assert named_range.destinations == ("Data!A2:A3",)
    assert named_range.status == "resolved"
    assert named_range.diagnostics == ()


def test_extract_workbook_classifies_ref_defined_name_as_source_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "ref-defined-name.xlsx"
    source = Workbook()
    source.active.title = "Inputs"
    source.active["A1"] = 1
    source.defined_names.add(DefinedName("BrokenName", attr_text="#REF!"))
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)
    named_range = workbook.named_ranges[0]

    assert named_range.name == "BrokenName"
    assert named_range.destinations == ()
    assert named_range.status == "unresolved"
    assert named_range.diagnostics[0].code == "named_range_source_error"


def test_extract_workbook_reports_unresolved_non_range_defined_name(tmp_path: Path) -> None:
    workbook_path = tmp_path / "defined-name-formula.xlsx"
    source = Workbook()
    source.active.title = "Inputs"
    source.active["A1"] = 1
    source.defined_names.add(DefinedName("ComputedName", attr_text="=1+1"))
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)
    named_range = workbook.named_ranges[0]

    assert named_range.name == "ComputedName"
    assert named_range.destinations == ()
    assert named_range.status == "unresolved"
    assert named_range.diagnostics[0].code == "unresolved_named_range"


def test_extract_workbook_reads_sparse_populated_cells_without_dense_scan(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sparse.xlsx"
    source = Workbook()
    source.active.title = "Inputs"
    source.active["A1"] = 1
    source.active["CV5000"] = "=A1+1"
    source.save(workbook_path)

    workbook = extract_workbook(workbook_path)

    assert [cell.cell_ref for cell in workbook.cells] == ["Inputs!A1", "Inputs!CV5000"]
