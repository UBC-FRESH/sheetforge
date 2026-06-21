"""Generated Python module contracts."""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl.utils.cell import get_column_letter, range_boundaries

from sheetforge.extraction import WorkbookRecord
from sheetforge.formulas import FormulaExpression, FormulaExpressionNode
from sheetforge.graph import DependencyGraph


JsonValue = str | int | float | bool | None | list[Any] | dict[str, Any]
DiagnosticSeverity = Literal["info", "warning", "error"]
GeneratedSymbolKind = Literal["input", "intermediate", "output"]


@dataclass(frozen=True)
class GenerationDiagnostic:
    """Generation concern tied to workbook or generated-code provenance."""

    code: str
    message: str
    severity: DiagnosticSeverity = "warning"
    location: str | None = None
    raw_value: JsonValue = None

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GenerationDiagnostic":
        return cls(
            code=data["code"],
            message=data["message"],
            severity=data.get("severity", "warning"),
            location=data.get("location"),
            raw_value=data.get("raw_value"),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
            "location": self.location,
            "raw_value": self.raw_value,
        }


@dataclass(frozen=True)
class GeneratedSymbol:
    """Generated Python symbol tied back to workbook provenance."""

    cell_ref: str
    symbol_name: str
    kind: GeneratedSymbolKind
    raw_formula: str | None = None

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GeneratedSymbol":
        return cls(
            cell_ref=data["cell_ref"],
            symbol_name=data["symbol_name"],
            kind=data["kind"],
            raw_formula=data.get("raw_formula"),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "cell_ref": self.cell_ref,
            "symbol_name": self.symbol_name,
            "kind": self.kind,
            "raw_formula": self.raw_formula,
        }


@dataclass(frozen=True)
class GeneratedModuleContract:
    """Contract for one generated standalone Python module."""

    workbook_id: str
    module_name: str
    entrypoint: str = "calculate"
    input_refs: tuple[str, ...] = field(default_factory=tuple)
    output_refs: tuple[str, ...] = field(default_factory=tuple)
    symbols: tuple[GeneratedSymbol, ...] = field(default_factory=tuple)
    include_provenance_comments: bool = True

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GeneratedModuleContract":
        return cls(
            workbook_id=data["workbook_id"],
            module_name=data["module_name"],
            entrypoint=data.get("entrypoint", "calculate"),
            input_refs=tuple(data.get("input_refs", [])),
            output_refs=tuple(data.get("output_refs", [])),
            symbols=tuple(GeneratedSymbol.from_dict(item) for item in data.get("symbols", [])),
            include_provenance_comments=data.get("include_provenance_comments", True),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "workbook_id": self.workbook_id,
            "module_name": self.module_name,
            "entrypoint": self.entrypoint,
            "input_refs": list(self.input_refs),
            "output_refs": list(self.output_refs),
            "symbols": [symbol.to_dict() for symbol in self.symbols],
            "include_provenance_comments": self.include_provenance_comments,
        }


@dataclass(frozen=True)
class GenerationResult:
    """Result of generating one Python module."""

    contract: GeneratedModuleContract
    source_code: str = ""
    diagnostics: tuple[GenerationDiagnostic, ...] = field(default_factory=tuple)

    @property
    def generated(self) -> bool:
        return bool(self.source_code) and not any(diagnostic.severity == "error" for diagnostic in self.diagnostics)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GenerationResult":
        return cls(
            contract=GeneratedModuleContract.from_dict(data["contract"]),
            source_code=data.get("source_code", ""),
            diagnostics=tuple(GenerationDiagnostic.from_dict(item) for item in data.get("diagnostics", [])),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "contract": self.contract.to_dict(),
            "source_code": self.source_code,
            "generated": self.generated,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


@dataclass(frozen=True)
class GeneratedContractInferenceResult:
    """Generated-model contract inference result for selected workbook outputs."""

    contract: GeneratedModuleContract
    expressions: dict[str, FormulaExpression] = field(default_factory=dict)
    constants: dict[str, JsonValue] = field(default_factory=dict)
    diagnostics: tuple[GenerationDiagnostic, ...] = field(default_factory=tuple)

    @property
    def inferred(self) -> bool:
        return not any(diagnostic.severity == "error" for diagnostic in self.diagnostics)

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "contract": self.contract.to_dict(),
            "expressions": {cell_ref: expression.to_dict() for cell_ref, expression in self.expressions.items()},
            "constants": self.constants,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
            "inferred": self.inferred,
        }


def symbol_name_for_cell_ref(cell_ref: str) -> str:
    """Build a stable Python identifier from a canonical workbook cell ref."""

    symbol = re.sub(r"[^0-9A-Za-z_]+", "_", cell_ref).strip("_").lower()
    if not symbol:
        return "cell"
    if symbol[0].isdigit():
        return f"cell_{symbol}"
    return symbol


def infer_generated_module_contract(
    *,
    workbook: WorkbookRecord,
    graph: DependencyGraph,
    expressions: Mapping[str, FormulaExpression],
    output_refs: Sequence[str],
    module_name: str,
    input_refs: Sequence[str] = (),
) -> GeneratedContractInferenceResult:
    """Infer a generated module contract by walking dependencies for selected outputs."""

    explicit_inputs = set(input_refs)
    selected_outputs = tuple(output_refs)
    cell_by_ref = {cell.cell_ref: cell for cell in workbook.cells}
    dependencies_by_target: dict[str, list[str]] = {}
    edge_diagnostics_by_target: dict[str, list[GenerationDiagnostic]] = {}
    diagnostics: list[GenerationDiagnostic] = []

    for edge in graph.execution_edges:
        if edge.diagnostic_code is not None:
            edge_diagnostics_by_target.setdefault(edge.target.normalized, []).append(
                GenerationDiagnostic(
                    code="unsupported_dependency_edge",
                    message="dependency edge has a diagnostic and cannot be inferred silently",
                    severity="error",
                    location=edge.target.normalized,
                    raw_value=edge.diagnostic_code,
                )
            )
            continue
        if edge.source.kind != "cell":
            edge_diagnostics_by_target.setdefault(edge.target.normalized, []).append(
                GenerationDiagnostic(
                    code="unsupported_dependency_source",
                    message="dependency source is not a concrete cell reference",
                    severity="error",
                    location=edge.target.normalized,
                    raw_value=edge.source.normalized,
                )
            )
            continue
        dependencies_by_target.setdefault(edge.target.normalized, []).append(edge.source.normalized)

    input_order: list[str] = []
    formula_order: list[str] = []
    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(cell_ref: str) -> None:
        if cell_ref in visited:
            return
        if cell_ref in visiting:
            diagnostics.append(
                GenerationDiagnostic(
                    code="circular_dependency",
                    message="selected output dependency walk encountered a cycle",
                    severity="error",
                    location=cell_ref,
                )
            )
            return

        cell = cell_by_ref.get(cell_ref)
        if cell is None:
            diagnostics.append(
                GenerationDiagnostic(
                    code="missing_dependency_cell",
                    message="selected output depends on a cell that was not extracted",
                    severity="error",
                    location=cell_ref,
                )
            )
            return

        if cell_ref in explicit_inputs or cell.formula is None:
            if cell_ref not in input_order:
                input_order.append(cell_ref)
            visited.add(cell_ref)
            return

        visiting.add(cell_ref)
        diagnostics.extend(edge_diagnostics_by_target.get(cell_ref, ()))
        for dependency_ref in dependencies_by_target.get(cell_ref, []):
            visit(dependency_ref)
        visiting.remove(cell_ref)

        if cell_ref not in formula_order:
            formula_order.append(cell_ref)
        visited.add(cell_ref)

    for output_ref in selected_outputs:
        visit(output_ref)

    selected_expressions: dict[str, FormulaExpression] = {}
    for cell_ref in formula_order:
        expression = expressions.get(cell_ref)
        if expression is None:
            diagnostics.append(
                GenerationDiagnostic(
                    code="missing_formula_expression",
                    message="inferred generated symbol has no translated formula expression",
                    severity="error",
                    location=cell_ref,
                )
            )
            continue
        selected_expressions[cell_ref] = expression

    constants = {cell_ref: cell_by_ref[cell_ref].raw_value for cell_ref in input_order if cell_ref in cell_by_ref}
    output_set = set(selected_outputs)
    symbols = tuple(
        GeneratedSymbol(cell_ref=cell_ref, symbol_name=symbol_name_for_cell_ref(cell_ref), kind="input")
        for cell_ref in input_order
    ) + tuple(
        GeneratedSymbol(
            cell_ref=cell_ref,
            symbol_name=symbol_name_for_cell_ref(cell_ref),
            kind="output" if cell_ref in output_set else "intermediate",
            raw_formula=cell_by_ref[cell_ref].formula.raw_formula if cell_by_ref[cell_ref].formula else None,
        )
        for cell_ref in formula_order
    )
    contract = GeneratedModuleContract(
        workbook_id=workbook.workbook_id,
        module_name=module_name,
        input_refs=tuple(input_order),
        output_refs=selected_outputs,
        symbols=symbols,
    )
    return GeneratedContractInferenceResult(
        contract=contract,
        expressions=selected_expressions,
        constants=constants,
        diagnostics=tuple(diagnostics),
    )


def generate_python_module(
    *,
    contract: GeneratedModuleContract,
    expressions: Mapping[str, FormulaExpression],
    constants: Mapping[str, JsonValue] | None = None,
    output_path: str | Path | None = None,
) -> GenerationResult:
    """Generate standalone Python source from translated formula expressions."""

    constants = constants or {}
    diagnostics = _generation_diagnostics(contract, expressions)
    if any(diagnostic.severity == "error" for diagnostic in diagnostics):
        return GenerationResult(contract=contract, diagnostics=tuple(diagnostics))

    source_code = _render_module(contract=contract, expressions=expressions, constants=constants)
    if output_path is not None:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(source_code, encoding="utf-8")

    return GenerationResult(contract=contract, source_code=source_code, diagnostics=tuple(diagnostics))


def _generation_diagnostics(
    contract: GeneratedModuleContract,
    expressions: Mapping[str, FormulaExpression],
) -> list[GenerationDiagnostic]:
    diagnostics: list[GenerationDiagnostic] = []
    for symbol in contract.symbols:
        if symbol.kind == "input":
            continue

        expression = expressions.get(symbol.cell_ref)
        if expression is None:
            diagnostics.append(
                GenerationDiagnostic(
                    code="missing_formula_expression",
                    message="generated symbol has no translated formula expression",
                    severity="error",
                    location=symbol.cell_ref,
                    raw_value=symbol.raw_formula,
                )
            )
            continue

        if not expression.translated:
            diagnostics.append(
                GenerationDiagnostic(
                    code="unsupported_formula",
                    message="formula expression could not be generated",
                    severity="error",
                    location=symbol.cell_ref,
                    raw_value=expression.raw_formula,
                )
            )
    return diagnostics


def _render_module(
    *,
    contract: GeneratedModuleContract,
    expressions: Mapping[str, FormulaExpression],
    constants: Mapping[str, JsonValue],
) -> str:
    lines = [
        '"""Generated Sheetforge model.',
        "",
        f"Source workbook: {contract.workbook_id}",
        '"""',
        "",
        "import fnmatch",
        "",
        "",
        "def _sf_flatten(values):",
        "    for value in values:",
        "        if isinstance(value, (list, tuple)):",
        "            yield from _sf_flatten(value)",
        "        else:",
        "            yield value",
        "",
        "",
        "def _sf_average(values):",
        "    values = list(values)",
        "    return sum(values) / len(values)",
        "",
        "",
        "def _sf_iferror(value_fn, fallback):",
        "    try:",
        "        return value_fn()",
        "    except Exception:",
        "        return fallback",
        "",
        "",
        "def _sf_ifna(value_fn, fallback):",
        "    try:",
        "        value = value_fn()",
        "    except LookupError:",
        "        return fallback",
        "    if value == '#N/A':",
        "        return fallback",
        "    return value",
        "",
        "",
        "def _sf_coerce_criteria(raw, sample):",
        "    if isinstance(raw, str):",
        "        upper = raw.upper()",
        "        if upper == 'TRUE':",
        "            return True",
        "        if upper == 'FALSE':",
        "            return False",
        "        try:",
        "            number = float(raw)",
        "        except ValueError:",
        "            return raw",
        "        if number.is_integer():",
        "            return int(number)",
        "        return number",
        "    return raw",
        "",
        "",
        "def _sf_compare_criteria(value, operator, expected):",
        "    if operator == '=':",
        "        return value == expected",
        "    if operator == '<>':",
        "        return value != expected",
        "    if operator == '>':",
        "        return value > expected",
        "    if operator == '>=':",
        "        return value >= expected",
        "    if operator == '<':",
        "        return value < expected",
        "    if operator == '<=':",
        "        return value <= expected",
        "    raise ValueError(f'unsupported criteria operator: {operator}')",
        "",
        "",
        "def _sf_matches_criteria(value, criteria):",
        "    if isinstance(criteria, str):",
        "        for operator in ('>=', '<=', '<>', '>', '<', '='):",
        "            if criteria.startswith(operator):",
        "                expected = _sf_coerce_criteria(criteria[len(operator):], value)",
        "                return _sf_compare_criteria(value, operator, expected)",
        "        if '*' in criteria or '?' in criteria:",
        "            return fnmatch.fnmatchcase(str(value), criteria)",
        "    return value == criteria",
        "",
        "",
        "def _sf_sumif(criteria_range, criteria, sum_range=None):",
        "    criteria_values = tuple(_sf_flatten((criteria_range,)))",
        "    sum_values = criteria_values if sum_range is None else tuple(_sf_flatten((sum_range,)))",
        "    return sum(",
        "        sum_value",
        "        for criteria_value, sum_value in zip(criteria_values, sum_values)",
        "        if _sf_matches_criteria(criteria_value, criteria)",
        "    )",
        "",
        "",
        "def _sf_countif(criteria_range, criteria):",
        "    return sum(1 for value in _sf_flatten((criteria_range,)) if _sf_matches_criteria(value, criteria))",
        "",
        "",
        "def _sf_sumifs(sum_range, *criteria_pairs):",
        "    sum_values = tuple(_sf_flatten((sum_range,)))",
        "    criteria_ranges = [tuple(_sf_flatten((criteria_range,))) for criteria_range, _criteria in criteria_pairs]",
        "    total = 0",
        "    for index, sum_value in enumerate(sum_values):",
        "        if all(_sf_matches_criteria(criteria_range[index], criteria) for criteria_range, criteria in zip(criteria_ranges, (criteria for _range, criteria in criteria_pairs))):",
        "            total += sum_value",
        "    return total",
        "",
        "",
        "def _sf_countifs(*criteria_pairs):",
        "    criteria_ranges = [tuple(_sf_flatten((criteria_range,))) for criteria_range, _criteria in criteria_pairs]",
        "    if not criteria_ranges:",
        "        return 0",
        "    criteria_values = tuple(criteria for _range, criteria in criteria_pairs)",
        "    return sum(",
        "        1",
        "        for index in range(len(criteria_ranges[0]))",
        "        if all(_sf_matches_criteria(criteria_range[index], criteria) for criteria_range, criteria in zip(criteria_ranges, criteria_values))",
        "    )",
        "",
        "",
        "def _sf_range_lookup_enabled(range_lookup):",
        "    if isinstance(range_lookup, str):",
        "        return range_lookup.upper() not in {'FALSE', '0'}",
        "    return bool(range_lookup)",
        "",
        "",
        "def _sf_vlookup(lookup_value, table_array, col_index_num, range_lookup=True):",
        "    column_index = int(col_index_num) - 1",
        "    if column_index < 0:",
        "        raise ValueError('VLOOKUP column index must be one-based')",
        "    rows = tuple(tuple(row) for row in table_array)",
        "    if not rows:",
        "        raise LookupError('VLOOKUP table is empty')",
        "    if any(column_index >= len(row) for row in rows):",
        "        raise IndexError('VLOOKUP column index is outside the table')",
        "    if not _sf_range_lookup_enabled(range_lookup):",
        "        for row in rows:",
        "            if row[0] == lookup_value:",
        "                return row[column_index]",
        "        raise LookupError('VLOOKUP exact match not found')",
        "    candidate = None",
        "    for row in rows:",
        "        try:",
        "            matched = row[0] <= lookup_value",
        "        except TypeError:",
        "            continue",
        "        if matched:",
        "            candidate = row",
        "        else:",
        "            break",
        "    if candidate is None:",
        "        raise LookupError('VLOOKUP approximate match not found')",
        "    return candidate[column_index]",
        "",
        "",
        f"def {contract.entrypoint}(inputs=None):",
        "    inputs = {} if inputs is None else dict(inputs)",
    ]

    for symbol in contract.symbols:
        if contract.include_provenance_comments:
            lines.append(f"    # {symbol.cell_ref}" + (f": {symbol.raw_formula}" if symbol.raw_formula else ""))

        if symbol.kind == "input":
            default_value = constants.get(symbol.cell_ref)
            lines.append(
                f"    {symbol.symbol_name} = inputs.get({symbol.cell_ref!r}, {default_value!r})"
            )
            continue

        expression = expressions[symbol.cell_ref]
        lines.append(f"    {symbol.symbol_name} = {_render_expression(expression.root)}")

    lines.append("    return {")
    for output_ref in contract.output_refs:
        lines.append(f"        {output_ref!r}: {symbol_name_for_cell_ref(output_ref)},")
    lines.append("    }")
    lines.append("")
    return "\n".join(lines)


def _render_expression(node: FormulaExpressionNode | None) -> str:
    if node is None:
        raise ValueError("cannot render missing formula expression root")

    if node.kind == "literal":
        return repr(node.value)
    if node.kind == "reference":
        if node.reference is None:
            raise ValueError("cannot render reference expression without reference")
        if node.reference.kind == "range":
            return _render_range_reference(node.reference)
        return symbol_name_for_cell_ref(node.reference.normalized)
    if node.kind == "unary":
        (operand,) = node.operands
        if node.operator == "-":
            return f"(-{_render_expression(operand)})"
        raise ValueError(f"unsupported unary operator: {node.operator}")
    if node.kind == "binary":
        left, right = node.operands
        if node.operator == "^":
            return f"({_render_expression(left)} ** {_render_expression(right)})"
        if node.operator == "&":
            return f"(str({_render_expression(left)}) + str({_render_expression(right)}))"
        return f"({_render_expression(left)} {node.operator} {_render_expression(right)})"
    if node.kind == "comparison":
        left, right = node.operands
        operator = _python_comparison_operator(node.operator)
        return f"({_render_expression(left)} {operator} {_render_expression(right)})"
    if node.kind == "function_call":
        return _render_function_call(node)

    raise ValueError(f"unsupported expression kind: {node.kind}")


def _render_function_call(node: FormulaExpressionNode) -> str:
    if node.function_name == "ROUND":
        if len(node.operands) != 2:
            raise ValueError("ROUND requires two operands")
        return f"round({_render_expression(node.operands[0])}, {_render_expression(node.operands[1])})"
    if node.function_name == "IF":
        if len(node.operands) != 3:
            raise ValueError("IF requires three operands")
        condition, true_value, false_value = node.operands
        return f"({_render_expression(true_value)} if {_render_expression(condition)} else {_render_expression(false_value)})"
    if node.function_name == "IFERROR":
        if len(node.operands) != 2:
            raise ValueError("IFERROR requires two operands")
        value, fallback = node.operands
        return f"_sf_iferror(lambda: {_render_expression(value)}, {_render_expression(fallback)})"
    if node.function_name == "IFNA":
        if len(node.operands) != 2:
            raise ValueError("IFNA requires two operands")
        value, fallback = node.operands
        return f"_sf_ifna(lambda: {_render_expression(value)}, {_render_expression(fallback)})"
    if node.function_name == "AND":
        return f"all(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "OR":
        return f"any(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "SUM":
        return f"sum(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "MIN":
        return f"min(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "MAX":
        return f"max(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "AVERAGE":
        return f"_sf_average(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "CONCATENATE":
        return f"''.join(str(value) for value in _sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "SUMIF":
        if len(node.operands) not in {2, 3}:
            raise ValueError("SUMIF requires two or three operands")
        return f"_sf_sumif({_render_function_arguments(node.operands)})"
    if node.function_name == "COUNTIF":
        if len(node.operands) != 2:
            raise ValueError("COUNTIF requires two operands")
        return f"_sf_countif({_render_function_arguments(node.operands)})"
    if node.function_name == "SUMIFS":
        if len(node.operands) < 3 or len(node.operands) % 2 != 1:
            raise ValueError("SUMIFS requires a sum range followed by criteria range/criteria pairs")
        return f"_sf_sumifs({_render_criteria_function_arguments(node.operands)})"
    if node.function_name == "COUNTIFS":
        if len(node.operands) < 2 or len(node.operands) % 2 != 0:
            raise ValueError("COUNTIFS requires criteria range/criteria pairs")
        return f"_sf_countifs({_render_criteria_function_arguments(node.operands)})"
    if node.function_name == "VLOOKUP":
        if len(node.operands) not in {3, 4}:
            raise ValueError("VLOOKUP requires three or four operands")
        lookup_value, table_array, col_index_num, *range_lookup = node.operands
        rendered_arguments = [
            _render_expression(lookup_value),
            _render_table_array(table_array),
            _render_expression(col_index_num),
        ]
        if range_lookup:
            rendered_arguments.append(_render_expression(range_lookup[0]))
        return f"_sf_vlookup({', '.join(rendered_arguments)})"
    raise ValueError(f"unsupported function call: {node.function_name}")


def _render_function_arguments(operands: tuple[FormulaExpressionNode, ...]) -> str:
    return ", ".join(_render_expression(operand) for operand in operands)


def _render_criteria_function_arguments(operands: tuple[FormulaExpressionNode, ...]) -> str:
    if len(operands) < 2:
        return ""

    rendered: list[str] = []
    if len(operands) % 2 == 1:
        rendered.append(_render_expression(operands[0]))
        pair_operands = operands[1:]
    else:
        pair_operands = operands

    rendered.extend(
        f"({_render_expression(pair_operands[index])}, {_render_expression(pair_operands[index + 1])})"
        for index in range(0, len(pair_operands), 2)
    )
    return ", ".join(rendered)


def _render_argument_tuple(operands: tuple[FormulaExpressionNode, ...]) -> str:
    rendered = ", ".join(_render_expression(operand) for operand in operands)
    if len(operands) == 1:
        rendered = f"{rendered},"
    return f"({rendered})"


def _render_range_reference(reference) -> str:
    if reference.sheet is None or reference.start_cell is None or reference.end_cell is None:
        raise ValueError(f"cannot render incomplete range reference: {reference.normalized}")

    min_col, min_row, max_col, max_row = range_boundaries(f"{reference.start_cell}:{reference.end_cell}")
    rendered_cells = [
        symbol_name_for_cell_ref(f"{reference.sheet}!{get_column_letter(column)}{row}")
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    ]
    return f"({', '.join(rendered_cells)}{',' if len(rendered_cells) == 1 else ''})"


def _render_table_array(node: FormulaExpressionNode) -> str:
    if node.kind != "reference" or node.reference is None or node.reference.kind != "range":
        raise ValueError("VLOOKUP table array must be a concrete range reference")
    reference = node.reference
    if reference.sheet is None or reference.start_cell is None or reference.end_cell is None:
        raise ValueError(f"cannot render incomplete VLOOKUP table reference: {reference.normalized}")

    min_col, min_row, max_col, max_row = range_boundaries(f"{reference.start_cell}:{reference.end_cell}")
    rendered_rows = []
    for row in range(min_row, max_row + 1):
        rendered_cells = [
            symbol_name_for_cell_ref(f"{reference.sheet}!{get_column_letter(column)}{row}")
            for column in range(min_col, max_col + 1)
        ]
        rendered_rows.append(f"({', '.join(rendered_cells)}{',' if len(rendered_cells) == 1 else ''})")
    return f"({', '.join(rendered_rows)}{',' if len(rendered_rows) == 1 else ''})"


def _python_comparison_operator(operator: str | None) -> str:
    if operator == "=":
        return "=="
    if operator == "<>":
        return "!="
    if operator is None:
        raise ValueError("missing comparison operator")
    return operator
