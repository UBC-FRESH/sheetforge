"""Workbook extraction records.

These records describe extracted workbook facts; they do not read workbook
files themselves.
"""

from __future__ import annotations

from collections.abc import Callable
from datetime import date, datetime, time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import get_column_letter, range_boundaries


JsonValue = str | int | float | bool | None | list[Any] | dict[str, Any]
CellKind = Literal["value", "formula", "blank", "error"]
DiagnosticSeverity = Literal["info", "warning", "error"]
NamedRangeStatus = Literal["resolved", "partially_resolved", "unresolved"]

VOLATILE_FUNCTIONS = frozenset({"NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT"})


@dataclass(frozen=True)
class ExtractionDiagnostic:
    """Extraction or interpretation concern tied to workbook provenance."""

    code: str
    message: str
    severity: DiagnosticSeverity = "warning"
    location: str | None = None
    raw_value: JsonValue = None

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "ExtractionDiagnostic":
        return cls(
            code=data["code"],
            message=data["message"],
            severity=data.get("severity", "warning"),
            location=data.get("location"),
            raw_value=data.get("raw_value"),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
            "location": self.location,
            "raw_value": self.raw_value,
        }


@dataclass(frozen=True)
class FormulaRecord:
    """Formula text and parsed extraction facts for one formula cell."""

    raw_formula: str
    tokens: tuple[str, ...] = field(default_factory=tuple)
    raw_references: tuple[str, ...] = field(default_factory=tuple)
    normalized_references: tuple[str, ...] = field(default_factory=tuple)
    functions: tuple[str, ...] = field(default_factory=tuple)
    diagnostics: tuple[ExtractionDiagnostic, ...] = field(default_factory=tuple)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "FormulaRecord":
        return cls(
            raw_formula=data["raw_formula"],
            tokens=tuple(data.get("tokens", [])),
            raw_references=tuple(data.get("raw_references", [])),
            normalized_references=tuple(data.get("normalized_references", [])),
            functions=tuple(data.get("functions", [])),
            diagnostics=tuple(ExtractionDiagnostic.from_dict(item) for item in data.get("diagnostics", [])),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "raw_formula": self.raw_formula,
            "tokens": list(self.tokens),
            "raw_references": list(self.raw_references),
            "normalized_references": list(self.normalized_references),
            "functions": list(self.functions),
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


@dataclass(frozen=True)
class CellRecord:
    """Extracted cell facts for one canonical workbook cell reference."""

    cell_ref: str
    kind: CellKind
    raw_value: JsonValue
    data_type: str | None = None
    cached_value: JsonValue = None
    formula: FormulaRecord | None = None

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "CellRecord":
        formula_data = data.get("formula")
        return cls(
            cell_ref=data["cell_ref"],
            kind=data["kind"],
            raw_value=data.get("raw_value"),
            data_type=data.get("data_type"),
            cached_value=data.get("cached_value"),
            formula=FormulaRecord.from_dict(formula_data) if formula_data is not None else None,
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "cell_ref": self.cell_ref,
            "kind": self.kind,
            "raw_value": self.raw_value,
            "data_type": self.data_type,
            "cached_value": self.cached_value,
            "formula": self.formula.to_dict() if self.formula is not None else None,
        }


@dataclass(frozen=True)
class NamedRangeRecord:
    """Workbook or worksheet scoped defined name."""

    name: str
    scope: str
    raw_definition: str
    destinations: tuple[str, ...] = field(default_factory=tuple)
    status: NamedRangeStatus = "unresolved"
    diagnostics: tuple[ExtractionDiagnostic, ...] = field(default_factory=tuple)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "NamedRangeRecord":
        return cls(
            name=data["name"],
            scope=data["scope"],
            raw_definition=data["raw_definition"],
            destinations=tuple(data.get("destinations", [])),
            status=data.get("status", "unresolved"),
            diagnostics=tuple(ExtractionDiagnostic.from_dict(item) for item in data.get("diagnostics", [])),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "name": self.name,
            "scope": self.scope,
            "raw_definition": self.raw_definition,
            "destinations": list(self.destinations),
            "status": self.status,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


@dataclass(frozen=True)
class TableRecord:
    """Worksheet table metadata needed to resolve structured references."""

    name: str
    sheet: str
    ref: str
    columns: tuple[str, ...] = field(default_factory=tuple)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "TableRecord":
        return cls(
            name=data["name"],
            sheet=data["sheet"],
            ref=data["ref"],
            columns=tuple(data.get("columns", [])),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "name": self.name,
            "sheet": self.sheet,
            "ref": self.ref,
            "columns": list(self.columns),
        }


@dataclass(frozen=True)
class SheetRecord:
    """Worksheet identity and ordering facts."""

    sheet_id: str
    title: str
    state: str
    index: int

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "SheetRecord":
        return cls(
            sheet_id=data["sheet_id"],
            title=data["title"],
            state=data["state"],
            index=data["index"],
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "sheet_id": self.sheet_id,
            "title": self.title,
            "state": self.state,
            "index": self.index,
        }


@dataclass(frozen=True)
class WorkbookRecord:
    """Extracted facts for one source workbook."""

    workbook_id: str
    source_path: str
    sheets: tuple[SheetRecord, ...] = field(default_factory=tuple)
    cells: tuple[CellRecord, ...] = field(default_factory=tuple)
    named_ranges: tuple[NamedRangeRecord, ...] = field(default_factory=tuple)
    tables: tuple[TableRecord, ...] = field(default_factory=tuple)
    diagnostics: tuple[ExtractionDiagnostic, ...] = field(default_factory=tuple)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "WorkbookRecord":
        return cls(
            workbook_id=data["workbook_id"],
            source_path=data["source_path"],
            sheets=tuple(SheetRecord.from_dict(item) for item in data.get("sheets", [])),
            cells=tuple(CellRecord.from_dict(item) for item in data.get("cells", [])),
            named_ranges=tuple(NamedRangeRecord.from_dict(item) for item in data.get("named_ranges", [])),
            tables=tuple(TableRecord.from_dict(item) for item in data.get("tables", [])),
            diagnostics=tuple(ExtractionDiagnostic.from_dict(item) for item in data.get("diagnostics", [])),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "workbook_id": self.workbook_id,
            "source_path": self.source_path,
            "sheets": [sheet.to_dict() for sheet in self.sheets],
            "cells": [cell.to_dict() for cell in self.cells],
            "named_ranges": [named_range.to_dict() for named_range in self.named_ranges],
            "tables": [table.to_dict() for table in self.tables],
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


def extract_workbook(path: str | Path, progress: Callable[[str], None] | None = None) -> WorkbookRecord:
    """Extract workbook facts with openpyxl into Sheetforge records."""

    workbook_path = Path(path)
    _progress(progress, "load_workbook formulas start")
    workbook = load_workbook(workbook_path, data_only=False)
    _progress(progress, "load_workbook formulas done")
    _progress(progress, "load_workbook cached_values start")
    cached_workbook = load_workbook(workbook_path, data_only=True)
    _progress(progress, "load_workbook cached_values done")

    _progress(progress, "workbook diagnostics start")
    diagnostics = _workbook_diagnostics(workbook)
    _progress(progress, "workbook diagnostics done")
    sheets = tuple(
        SheetRecord(
            sheet_id=worksheet.title,
            title=worksheet.title,
            state=worksheet.sheet_state,
            index=index,
        )
        for index, worksheet in enumerate(workbook.worksheets)
    )
    _progress(progress, f"sheets extracted count={len(sheets)}")
    _progress(progress, "tables start")
    tables = tuple(table for worksheet in workbook.worksheets for table in _extract_tables(worksheet))
    _progress(progress, f"tables done count={len(tables)}")
    _progress(progress, "named ranges start")
    named_ranges = tuple(
        _extract_named_range(name, defined_name, tables=tables) for name, defined_name in workbook.defined_names.items()
    )
    _progress(progress, f"named ranges done count={len(named_ranges)}")

    cell_records: list[CellRecord] = []
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        populated_cells = _populated_cells(worksheet)
        _progress(
            progress,
            f"sheet cells start index={index}/{len(workbook.worksheets)} populated={len(populated_cells)}",
        )
        sheet_cells = _extract_sheet_cells(
            worksheet,
            cached_workbook[worksheet.title],
            populated_cells=populated_cells,
        )
        cell_records.extend(sheet_cells)
        _progress(
            progress,
            f"sheet cells done index={index}/{len(workbook.worksheets)} extracted={len(sheet_cells)} total={len(cell_records)}",
        )
    cells = tuple(cell_records)
    _progress(progress, f"workbook extraction done cells={len(cells)}")

    return WorkbookRecord(
        workbook_id=workbook_path.name,
        source_path=str(workbook_path),
        sheets=sheets,
        cells=cells,
        named_ranges=named_ranges,
        tables=tables,
        diagnostics=diagnostics,
    )


def _workbook_diagnostics(workbook: Any) -> tuple[ExtractionDiagnostic, ...]:
    diagnostics: list[ExtractionDiagnostic] = []
    if getattr(workbook, "vba_archive", None) is not None:
        diagnostics.append(
            ExtractionDiagnostic(
                code="unsupported_macros",
                message="workbook contains macros, which are not extracted",
                severity="warning",
                location="workbook",
            )
        )
    if getattr(workbook, "_external_links", None):
        diagnostics.append(
            ExtractionDiagnostic(
                code="unsupported_external_link",
                message="workbook contains external links, which are not extracted",
                severity="warning",
                location="workbook",
            )
        )
    return tuple(diagnostics)


def _extract_named_range(name: str, defined_name: Any, *, tables: tuple[TableRecord, ...] = ()) -> NamedRangeRecord:
    diagnostics: tuple[ExtractionDiagnostic, ...] = ()
    try:
        destinations = tuple(_cell_ref(sheet_name, coordinate) for sheet_name, coordinate in defined_name.destinations)
    except Exception:
        destinations = ()
    if not destinations:
        structured_destination = _structured_defined_name_destination(str(defined_name.attr_text), tables)
        if structured_destination is not None:
            destinations = (structured_destination,)
    status: NamedRangeStatus = "resolved" if destinations else "unresolved"
    if not destinations:
        code = "named_range_source_error" if str(defined_name.attr_text).upper() == "#REF!" else "unresolved_named_range"
        message = (
            "named range definition contains a source workbook error"
            if code == "named_range_source_error"
            else "named range destinations could not be resolved"
        )
        diagnostics = (
            ExtractionDiagnostic(
                code=code,
                message=message,
                severity="warning",
                location=name,
                raw_value=defined_name.attr_text,
            ),
        )

    scope = "workbook"
    local_sheet_id = getattr(defined_name, "localSheetId", None)
    if local_sheet_id is not None:
        scope = f"sheet:{local_sheet_id}"

    return NamedRangeRecord(
        name=name,
        scope=scope,
        raw_definition=defined_name.attr_text,
        destinations=destinations,
        status=status,
        diagnostics=diagnostics,
    )


def _extract_tables(worksheet: Any) -> tuple[TableRecord, ...]:
    records: list[TableRecord] = []
    for table in worksheet.tables.values():
        try:
            min_col, header_row, max_col, _max_row = range_boundaries(table.ref)
        except ValueError:
            continue

        columns = tuple(
            str(worksheet[f"{get_column_letter(column)}{header_row}"].value)
            for column in range(min_col, max_col + 1)
        )
        records.append(
            TableRecord(
                name=table.displayName,
                sheet=worksheet.title,
                ref=table.ref.replace("$", ""),
                columns=columns,
            )
        )
    return tuple(records)


def _structured_defined_name_destination(definition: str, tables: tuple[TableRecord, ...]) -> str | None:
    parsed = _parse_structured_defined_name(definition)
    if parsed is None:
        return None

    table = next((candidate for candidate in tables if candidate.name == parsed.table_name), None)
    if table is None:
        return None

    try:
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    except ValueError:
        return None

    try:
        column_offset = table.columns.index(parsed.column)
    except ValueError:
        return None

    column_name = get_column_letter(min_col + column_offset)
    start_row = min_row if parsed.include_headers else min_row + 1
    end_row = max_row
    if start_row > end_row:
        return None
    return f"{table.sheet}!{column_name}{start_row}:{column_name}{end_row}"


@dataclass(frozen=True)
class _StructuredDefinedName:
    table_name: str
    column: str
    include_headers: bool = False


def _parse_structured_defined_name(definition: str) -> _StructuredDefinedName | None:
    if not _is_structured_reference(definition):
        return None

    table_name = definition.split("[", 1)[0]
    if not table_name:
        return None

    bracketed_parts = _bracketed_parts(definition)
    column = next(
        (
            _clean_structured_selector(part)
            for part in reversed(bracketed_parts)
            if not part.startswith("#")
        ),
        None,
    )
    if column is None:
        return None
    return _StructuredDefinedName(
        table_name=table_name,
        column=column,
        include_headers=any(part in {"#All", "#Headers"} for part in bracketed_parts),
    )


def _extract_sheet_cells(
    worksheet: Any,
    cached_worksheet: Any,
    *,
    populated_cells: tuple[Any, ...] | None = None,
) -> tuple[CellRecord, ...]:
    records: list[CellRecord] = []
    for cell in populated_cells if populated_cells is not None else _populated_cells(worksheet):
        if cell.value is None:
            continue

        cell_ref = _cell_ref(worksheet.title, cell.coordinate)
        cached_value = cached_worksheet[cell.coordinate].value
        if cell.data_type == "f":
            formula = _extract_formula(cell_ref, str(cell.value), cached_value)
            records.append(
                CellRecord(
                    cell_ref=cell_ref,
                    kind="formula",
                    raw_value=_json_value(cell.value),
                    data_type=cell.data_type,
                    cached_value=_json_value(cached_value),
                    formula=formula,
                )
            )
            continue

        records.append(
            CellRecord(
                cell_ref=cell_ref,
                kind="value",
                raw_value=_json_value(cell.value),
                data_type=cell.data_type,
                cached_value=_json_value(cached_value),
                formula=None,
            )
        )
    return tuple(records)


def _populated_cells(worksheet: Any) -> tuple[Any, ...]:
    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict):
        return tuple(cell for _, cell in sorted(cells.items()))

    return tuple(cell for row in worksheet.iter_rows() for cell in row)


def _progress(progress: Callable[[str], None] | None, message: str) -> None:
    if progress is not None:
        progress(message)

def _extract_formula(cell_ref: str, raw_formula: str, cached_value: JsonValue) -> FormulaRecord:
    try:
        tokenizer = Tokenizer(raw_formula)
    except Exception as error:
        return FormulaRecord(
            raw_formula=raw_formula,
            diagnostics=(
                ExtractionDiagnostic(
                    code="formula_tokenization_failed",
                    message=f"formula could not be tokenized: {error}",
                    severity="warning",
                    location=cell_ref,
                    raw_value=raw_formula,
                ),
            ),
        )

    tokens = tuple(token.value for token in tokenizer.items)
    raw_references = tuple(
        token.value for token in tokenizer.items if token.type == "OPERAND" and token.subtype == "RANGE"
    )
    functions = tuple(
        token.value[:-1].upper() for token in tokenizer.items if token.type == "FUNC" and token.subtype == "OPEN"
    )
    diagnostics = _formula_diagnostics(
        cell_ref=cell_ref,
        raw_formula=raw_formula,
        cached_value=cached_value,
        functions=functions,
        raw_references=raw_references,
    )

    return FormulaRecord(
        raw_formula=raw_formula,
        tokens=tokens,
        raw_references=raw_references,
        normalized_references=(),
        functions=functions,
        diagnostics=diagnostics,
    )


def _formula_diagnostics(
    *,
    cell_ref: str,
    raw_formula: str,
    cached_value: JsonValue,
    functions: tuple[str, ...],
    raw_references: tuple[str, ...],
) -> tuple[ExtractionDiagnostic, ...]:
    diagnostics: list[ExtractionDiagnostic] = []
    if cached_value is None:
        diagnostics.append(
            ExtractionDiagnostic(
                code="missing_cached_formula_value",
                message="formula cell has no cached value",
                severity="warning",
                location=cell_ref,
                raw_value=raw_formula,
            )
        )

    for function in functions:
        if function in VOLATILE_FUNCTIONS:
            diagnostics.append(
                ExtractionDiagnostic(
                    code="unsupported_volatile_function",
                    message=f"formula uses volatile function {function}",
                    severity="warning",
                    location=cell_ref,
                    raw_value=raw_formula,
                )
            )

    for reference in raw_references:
        if _is_external_reference(reference):
            diagnostics.append(
                ExtractionDiagnostic(
                    code="unsupported_external_link",
                    message="formula references an external workbook",
                    severity="warning",
                    location=cell_ref,
                    raw_value=reference,
                )
            )
            continue

        if _is_structured_reference(reference):
            diagnostics.append(
                ExtractionDiagnostic(
                    code="unsupported_structured_reference",
                    message="formula uses an Excel structured reference",
                    severity="warning",
                    location=cell_ref,
                    raw_value=reference,
                )
            )

    return tuple(diagnostics)


def _cell_ref(sheet_name: str, coordinate: str) -> str:
    return f"{sheet_name}!{coordinate.replace('$', '')}"


def _json_value(value: Any) -> JsonValue:
    if isinstance(value, str | int | float | bool) or value is None:
        return value
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    return str(value)


def _is_external_reference(reference: str) -> bool:
    return "[" in reference and "]" in reference and ("." in reference.split("]", 1)[0] or "!" in reference)


def _is_structured_reference(reference: str) -> bool:
    return "[" in reference and "]" in reference and not _is_external_reference(reference)


def _bracketed_parts(reference: str) -> tuple[str, ...]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    for character in reference:
        if character == "[":
            if depth > 0:
                current.append(character)
            depth += 1
            continue
        if character == "]":
            depth -= 1
            if depth == 0:
                part = "".join(current)
                current = []
                if part.startswith("[") and part.endswith("]"):
                    parts.extend(_bracketed_parts(part))
                elif part:
                    parts.append(part)
                continue
            current.append(character)
            continue
        if depth > 0:
            current.append(character)
    return tuple(parts)


def _clean_structured_selector(selector: str) -> str:
    return selector.removeprefix("@").replace("''", "'")
