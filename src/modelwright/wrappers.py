"""Lightweight facades for generated Modelwright models."""

from __future__ import annotations

from collections.abc import Callable, Mapping
from dataclasses import dataclass, field
from types import ModuleType
from typing import Any, Literal

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, range_boundaries

from modelwright.references import normalize_cell_reference, normalize_reference


CellRole = Literal["input", "output", "intermediate", "metadata"]
GeneratedCalculate = Callable[[dict[str, object] | None], dict[str, object]]


class WrapperDeclarationError(ValueError):
    """Raised when wrapper declarations are malformed or inconsistent."""


class WrapperExecutionError(RuntimeError):
    """Raised when a generated model cannot be executed through a facade."""


@dataclass(frozen=True)
class CellRef:
    """Declared facade metadata for one generated-model cell reference."""

    cell_ref: str
    name: str
    label: str | None = None
    role: CellRole = "metadata"
    unit: str | None = None
    description: str | None = None


@dataclass(frozen=True)
class TableRef:
    """Declared rectangular workbook-like table facade."""

    name: str
    sheet: str
    range_ref: str
    cell_refs: tuple[tuple[str, ...], ...]
    row_labels: tuple[str, ...]
    column_labels: tuple[str, ...]
    role: CellRole = "output"
    label: str | None = None
    description: str | None = None


@dataclass(frozen=True)
class ReportRef:
    """Declared named output bundle."""

    name: str
    cells: tuple[str, ...] = ()
    tables: tuple[str, ...] = ()
    label: str | None = None
    description: str | None = None


@dataclass(frozen=True)
class CellView:
    """Inspection payload for one cell in a facade scenario."""

    cell_ref: str
    name: str | None = None
    label: str | None = None
    role: CellRole | None = None
    unit: str | None = None
    description: str | None = None
    value: object = None
    has_value: bool = False

    def to_dict(self) -> dict[str, object]:
        return {
            "cell_ref": self.cell_ref,
            "name": self.name,
            "label": self.label,
            "role": self.role,
            "unit": self.unit,
            "description": self.description,
            "value": self.value,
            "has_value": self.has_value,
        }


@dataclass(frozen=True)
class TableView:
    """Calculated rectangular table payload."""

    name: str
    sheet: str
    range_ref: str
    rows: tuple[str, ...]
    columns: tuple[str, ...]
    cell_refs: tuple[tuple[str, ...], ...]
    values: tuple[tuple[object, ...], ...]
    label: str | None = None
    description: str | None = None

    def to_dict(self) -> dict[str, object]:
        return {
            "name": self.name,
            "sheet": self.sheet,
            "range_ref": self.range_ref,
            "rows": list(self.rows),
            "columns": list(self.columns),
            "cell_refs": [list(row) for row in self.cell_refs],
            "values": [list(row) for row in self.values],
            "label": self.label,
            "description": self.description,
        }


@dataclass(frozen=True)
class Scenario:
    """Immutable input override set for a facade calculation."""

    name: str = "default"
    _inputs: tuple[tuple[str, object], ...] = field(default_factory=tuple)

    @classmethod
    def from_inputs(cls, name: str = "default", inputs: Mapping[str, object] | None = None) -> "Scenario":
        if inputs is None:
            return cls(name=name)
        normalized = {normalize_full_cell_ref(cell_ref): value for cell_ref, value in inputs.items()}
        return cls(name=name, _inputs=tuple(normalized.items()))

    @property
    def inputs(self) -> dict[str, object]:
        return dict(self._inputs)

    def with_input(self, cell_ref: str, value: object) -> "Scenario":
        inputs = self.inputs
        inputs[normalize_full_cell_ref(cell_ref)] = value
        return Scenario.from_inputs(name=self.name, inputs=inputs)


def cell(
    cell_ref: str,
    *,
    name: str | None = None,
    label: str | None = None,
    role: CellRole = "metadata",
    unit: str | None = None,
    description: str | None = None,
) -> CellRef:
    """Declare one facade cell."""

    normalized = normalize_full_cell_ref(cell_ref)
    return CellRef(
        cell_ref=normalized,
        name=name or normalized,
        label=label,
        role=role,
        unit=unit,
        description=description,
    )


def table(
    name: str,
    *,
    sheet: str | None = None,
    range_ref: str,
    row_labels: tuple[str, ...] | list[str] | None = None,
    column_labels: tuple[str, ...] | list[str] | None = None,
    role: CellRole = "output",
    label: str | None = None,
    description: str | None = None,
) -> TableRef:
    """Declare one rectangular facade table."""

    table_sheet, coordinate = normalize_table_range(sheet=sheet, range_ref=range_ref)
    min_col, min_row, max_col, max_row = range_boundaries(coordinate)
    row_count = max_row - min_row + 1
    column_count = max_col - min_col + 1

    rows = tuple(row_labels) if row_labels is not None else tuple(str(row) for row in range(min_row, max_row + 1))
    columns = (
        tuple(column_labels)
        if column_labels is not None
        else tuple(get_column_letter(column) for column in range(min_col, max_col + 1))
    )
    if len(rows) != row_count:
        raise WrapperDeclarationError(
            f"table {name!r} declares {len(rows)} row labels for {row_count} table rows"
        )
    if len(columns) != column_count:
        raise WrapperDeclarationError(
            f"table {name!r} declares {len(columns)} column labels for {column_count} table columns"
        )

    cell_refs = tuple(
        tuple(f"{table_sheet}!{get_column_letter(column)}{row}" for column in range(min_col, max_col + 1))
        for row in range(min_row, max_row + 1)
    )
    normalized_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return TableRef(
        name=name,
        sheet=table_sheet,
        range_ref=normalized_range,
        cell_refs=cell_refs,
        row_labels=rows,
        column_labels=columns,
        role=role,
        label=label,
        description=description,
    )


def report(
    name: str,
    *,
    cells: tuple[str, ...] | list[str] = (),
    tables: tuple[str, ...] | list[str] = (),
    label: str | None = None,
    description: str | None = None,
) -> ReportRef:
    """Declare one structured report selection."""

    return ReportRef(
        name=name,
        cells=tuple(cells),
        tables=tuple(tables),
        label=label,
        description=description,
    )


class ModelFacade:
    """Analyst-facing facade around a generated Modelwright model."""

    def __init__(
        self,
        generated_model: ModuleType | GeneratedCalculate | object,
        *,
        cells: tuple[CellRef, ...] | list[CellRef] = (),
        tables: tuple[TableRef, ...] | list[TableRef] = (),
        reports: tuple[ReportRef, ...] | list[ReportRef] = (),
    ) -> None:
        self._calculate = generated_calculate(generated_model)
        self._cells = keyed_declarations(cells, kind="cell")
        self._cells_by_ref = {declaration.cell_ref: declaration for declaration in self._cells.values()}
        self._tables = keyed_declarations(tables, kind="table")
        self._reports = keyed_declarations(reports, kind="report")
        self._last_scenario: Scenario | None = None
        self._last_inputs: dict[str, object] = {}
        self._last_values: dict[str, object] | None = None
        self._validate_reports()

    @property
    def cells(self) -> dict[str, CellRef]:
        return dict(self._cells)

    @property
    def tables(self) -> dict[str, TableRef]:
        return dict(self._tables)

    @property
    def reports(self) -> dict[str, ReportRef]:
        return dict(self._reports)

    def inputs(self) -> dict[str, CellRef]:
        return {
            name: declaration
            for name, declaration in self._cells.items()
            if declaration.role == "input"
        }

    def outputs(self) -> dict[str, CellRef]:
        return {
            name: declaration
            for name, declaration in self._cells.items()
            if declaration.role == "output"
        }

    def scenario(self, name: str = "default", inputs: Mapping[str, object] | None = None) -> Scenario:
        return Scenario.from_inputs(name=name, inputs=inputs)

    def calculate(self, scenario: Scenario | None = None) -> dict[str, object]:
        scenario = scenario or Scenario()
        try:
            values = self._calculate(scenario.inputs)
        except Exception as error:  # noqa: BLE001
            raise WrapperExecutionError(f"generated model calculation failed: {error!r}") from error
        if not isinstance(values, dict):
            raise WrapperExecutionError("generated model calculate() did not return a dictionary")
        self._last_scenario = scenario
        self._last_inputs = scenario.inputs
        self._last_values = dict(values)
        return dict(values)

    def inspect(self, cell_ref: str, scenario: Scenario | None = None) -> CellView:
        normalized = normalize_full_cell_ref(cell_ref)
        values = self._values_for(scenario)
        declaration = self._cells_by_ref.get(normalized)
        has_value = normalized in values
        return CellView(
            cell_ref=normalized,
            name=declaration.name if declaration else None,
            label=declaration.label if declaration else None,
            role=declaration.role if declaration else None,
            unit=declaration.unit if declaration else None,
            description=declaration.description if declaration else None,
            value=values.get(normalized),
            has_value=has_value,
        )

    def table(self, name: str, scenario: Scenario | None = None) -> TableView:
        declaration = self._tables.get(name)
        if declaration is None:
            raise WrapperDeclarationError(f"unknown table declaration {name!r}")
        values = self._values_for(scenario)
        table_values = tuple(
            tuple(values.get(cell_ref) for cell_ref in row)
            for row in declaration.cell_refs
        )
        return TableView(
            name=declaration.name,
            sheet=declaration.sheet,
            range_ref=declaration.range_ref,
            rows=declaration.row_labels,
            columns=declaration.column_labels,
            cell_refs=declaration.cell_refs,
            values=table_values,
            label=declaration.label,
            description=declaration.description,
        )

    def report(self, name: str, scenario: Scenario | None = None) -> dict[str, object]:
        declaration = self._reports.get(name)
        if declaration is None:
            raise WrapperDeclarationError(f"unknown report declaration {name!r}")
        cell_payloads = {}
        for cell_name in declaration.cells:
            cell_declaration = self._cells.get(cell_name)
            if cell_declaration is None:
                raise WrapperDeclarationError(f"report {name!r} references unknown cell {cell_name!r}")
            cell_payloads[cell_name] = self.inspect(cell_declaration.cell_ref, scenario=scenario).to_dict()
        table_payloads = {
            table_name: self.table(table_name, scenario=scenario).to_dict()
            for table_name in declaration.tables
        }
        return {
            "name": declaration.name,
            "label": declaration.label,
            "description": declaration.description,
            "cells": cell_payloads,
            "tables": table_payloads,
        }

    def _values_for(self, scenario: Scenario | None) -> dict[str, object]:
        if scenario is None and self._last_values is not None:
            return {**self._last_inputs, **self._last_values}
        active_scenario = scenario or Scenario()
        values = self.calculate(active_scenario)
        return {**active_scenario.inputs, **values}

    def _validate_reports(self) -> None:
        for report_declaration in self._reports.values():
            for cell_name in report_declaration.cells:
                if cell_name not in self._cells:
                    raise WrapperDeclarationError(
                        f"report {report_declaration.name!r} references unknown cell {cell_name!r}"
                    )
            for table_name in report_declaration.tables:
                if table_name not in self._tables:
                    raise WrapperDeclarationError(
                        f"report {report_declaration.name!r} references unknown table {table_name!r}"
                    )


def generated_calculate(generated_model: ModuleType | GeneratedCalculate | object) -> GeneratedCalculate:
    if callable(generated_model):
        return generated_model
    calculate = getattr(generated_model, "calculate", None)
    if callable(calculate):
        return calculate
    raise WrapperDeclarationError("generated model must be callable or expose a callable calculate()")


def keyed_declarations(declarations: tuple[Any, ...] | list[Any], *, kind: str) -> dict[str, Any]:
    keyed: dict[str, Any] = {}
    for declaration in declarations:
        name = getattr(declaration, "name", None)
        if not isinstance(name, str) or not name:
            raise WrapperDeclarationError(f"{kind} declaration is missing a non-empty name")
        if name in keyed:
            raise WrapperDeclarationError(f"duplicate {kind} declaration {name!r}")
        keyed[name] = declaration
    return keyed


def normalize_full_cell_ref(cell_ref: str) -> str:
    normalized = normalize_cell_reference(cell_ref)
    if normalized.kind != "cell" or normalized.sheet is None:
        raise WrapperDeclarationError(f"expected a full cell reference like 'Sheet!A1', got {cell_ref!r}")
    return normalized.normalized


def normalize_table_range(*, sheet: str | None, range_ref: str) -> tuple[str, str]:
    if sheet is not None:
        reference = normalize_reference(f"{sheet}!{range_ref}")
    else:
        reference = normalize_reference(range_ref)
    if reference.kind != "range" or reference.sheet is None or reference.start_cell is None or reference.end_cell is None:
        raise WrapperDeclarationError(
            f"expected a rectangular range with a sheet, got sheet={sheet!r} range_ref={range_ref!r}"
        )
    start_col, start_row = split_cell(reference.start_cell)
    end_col, end_row = split_cell(reference.end_cell)
    return reference.sheet, f"{start_col}{start_row}:{end_col}{end_row}"


def split_cell(cell_ref: str) -> tuple[str, int]:
    column = "".join(character for character in cell_ref if character.isalpha())
    row = "".join(character for character in cell_ref if character.isdigit())
    if not column or not row:
        raise WrapperDeclarationError(f"invalid cell coordinate {cell_ref!r}")
    return get_column_letter(column_index_from_string(column)), int(row)
