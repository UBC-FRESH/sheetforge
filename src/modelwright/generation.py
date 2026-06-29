"""Generated Python module contracts."""

from __future__ import annotations

import re
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl.utils.cell import get_column_letter, range_boundaries

from modelwright.extraction import WorkbookRecord
from modelwright.formulas import FormulaExpression, FormulaExpressionNode
from modelwright.graph import DependencyGraph
from modelwright.references import WorkbookReference


JsonValue = str | int | float | bool | None | list[Any] | dict[str, Any]
DEFAULT_INLINE_PROVENANCE_COMMENT_LIMIT = 50_000
DEFAULT_INLINE_FORMULA_LAMBDA_LIMIT = 50_000
DiagnosticSeverity = Literal["info", "warning", "error"]
FormulaStorage = Literal["lambdas", "expression_source"]
GeneratedSymbolKind = Literal["input", "intermediate", "output"]


@dataclass(frozen=True)
class GenerationDiagnostic:
    """Generation concern tied to workbook or generated-code provenance."""

    code: str
    message: str
    severity: DiagnosticSeverity = "warning"
    location: str | None = None
    raw_value: JsonValue = None

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GenerationDiagnostic":
        return cls(
            code=data["code"],
            message=data["message"],
            severity=data.get("severity", "warning"),
            location=data.get("location"),
            raw_value=data.get("raw_value"),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
            "location": self.location,
            "raw_value": self.raw_value,
        }


@dataclass(frozen=True)
class GeneratedSymbol:
    """Generated Python symbol tied back to workbook provenance."""

    cell_ref: str
    symbol_name: str
    kind: GeneratedSymbolKind
    raw_formula: str | None = None

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GeneratedSymbol":
        return cls(
            cell_ref=data["cell_ref"],
            symbol_name=data["symbol_name"],
            kind=data["kind"],
            raw_formula=data.get("raw_formula"),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "cell_ref": self.cell_ref,
            "symbol_name": self.symbol_name,
            "kind": self.kind,
            "raw_formula": self.raw_formula,
        }


@dataclass(frozen=True)
class GeneratedModuleContract:
    """Contract for one generated standalone Python module."""

    workbook_id: str
    module_name: str
    entrypoint: str = "calculate"
    input_refs: tuple[str, ...] = field(default_factory=tuple)
    output_refs: tuple[str, ...] = field(default_factory=tuple)
    symbols: tuple[GeneratedSymbol, ...] = field(default_factory=tuple)
    include_provenance_comments: bool = True
    formula_storage: FormulaStorage = "lambdas"

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GeneratedModuleContract":
        return cls(
            workbook_id=data["workbook_id"],
            module_name=data["module_name"],
            entrypoint=data.get("entrypoint", "calculate"),
            input_refs=tuple(data.get("input_refs", [])),
            output_refs=tuple(data.get("output_refs", [])),
            symbols=tuple(GeneratedSymbol.from_dict(item) for item in data.get("symbols", [])),
            include_provenance_comments=data.get("include_provenance_comments", True),
            formula_storage=data.get("formula_storage", "lambdas"),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "workbook_id": self.workbook_id,
            "module_name": self.module_name,
            "entrypoint": self.entrypoint,
            "input_refs": list(self.input_refs),
            "output_refs": list(self.output_refs),
            "symbols": [symbol.to_dict() for symbol in self.symbols],
            "include_provenance_comments": self.include_provenance_comments,
            "formula_storage": self.formula_storage,
        }


@dataclass(frozen=True)
class GenerationResult:
    """Result of generating one Python module."""

    contract: GeneratedModuleContract
    source_code: str = ""
    diagnostics: tuple[GenerationDiagnostic, ...] = field(default_factory=tuple)

    @property
    def generated(self) -> bool:
        return bool(self.source_code) and not any(diagnostic.severity == "error" for diagnostic in self.diagnostics)

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "GenerationResult":
        return cls(
            contract=GeneratedModuleContract.from_dict(data["contract"]),
            source_code=data.get("source_code", ""),
            diagnostics=tuple(GenerationDiagnostic.from_dict(item) for item in data.get("diagnostics", [])),
        )

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "contract": self.contract.to_dict(),
            "source_code": self.source_code,
            "generated": self.generated,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


@dataclass(frozen=True)
class GeneratedContractInferenceResult:
    """Generated-model contract inference result for selected workbook outputs."""

    contract: GeneratedModuleContract
    expressions: dict[str, FormulaExpression] = field(default_factory=dict)
    constants: dict[str, JsonValue] = field(default_factory=dict)
    diagnostics: tuple[GenerationDiagnostic, ...] = field(default_factory=tuple)

    @property
    def inferred(self) -> bool:
        return not any(diagnostic.severity == "error" for diagnostic in self.diagnostics)

    def to_dict(self) -> dict[str, JsonValue]:
        return {
            "contract": self.contract.to_dict(),
            "expressions": {cell_ref: expression.to_dict() for cell_ref, expression in self.expressions.items()},
            "constants": self.constants,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
            "inferred": self.inferred,
        }


def symbol_name_for_cell_ref(cell_ref: str) -> str:
    """Build a stable Python identifier from a canonical workbook cell ref."""

    symbol = re.sub(r"[^0-9A-Za-z_]+", "_", cell_ref).strip("_").lower()
    if not symbol:
        return "cell"
    if symbol[0].isdigit():
        return f"cell_{symbol}"
    return symbol


def infer_generated_module_contract(
    *,
    workbook: WorkbookRecord,
    graph: DependencyGraph,
    expressions: Mapping[str, FormulaExpression],
    output_refs: Sequence[str],
    module_name: str,
    input_refs: Sequence[str] = (),
    progress: Callable[[str], None] | None = None,
    inline_provenance_comment_limit: int | None = DEFAULT_INLINE_PROVENANCE_COMMENT_LIMIT,
    inline_formula_lambda_limit: int | None = DEFAULT_INLINE_FORMULA_LAMBDA_LIMIT,
) -> GeneratedContractInferenceResult:
    """Infer a generated module contract by walking dependencies for selected outputs."""

    explicit_inputs = set(input_refs)
    selected_outputs = tuple(output_refs)
    _progress(progress, f"contract inference start outputs={len(selected_outputs)} expressions={len(expressions)}")
    cell_by_ref = {cell.cell_ref: cell for cell in workbook.cells}
    dependencies_by_target: dict[str, list[str | WorkbookReference]] = {}
    edge_diagnostics_by_target: dict[str, list[GenerationDiagnostic]] = {}
    diagnostics: list[GenerationDiagnostic] = []

    execution_edges = graph.execution_edges
    _progress(progress, f"contract inference edge scan start execution_edges={len(execution_edges)}")
    for index, edge in enumerate(execution_edges, start=1):
        if edge.diagnostic_code is not None:
            edge_diagnostics_by_target.setdefault(edge.target.normalized, []).append(
                GenerationDiagnostic(
                    code="unsupported_dependency_edge",
                    message="dependency edge has a diagnostic and cannot be inferred silently",
                    severity="error",
                    location=edge.target.normalized,
                    raw_value=edge.diagnostic_code,
                )
            )
            continue
        if edge.source.kind == "range":
            dependencies_by_target.setdefault(edge.target.normalized, []).append(edge.source)
            if index == 1 or index % 100000 == 0 or index == len(execution_edges):
                _progress(
                    progress,
                    f"contract inference edge scan progress execution_edges={index}/{len(execution_edges)}",
                )
            continue
        if edge.source.kind != "cell":
            edge_diagnostics_by_target.setdefault(edge.target.normalized, []).append(
                GenerationDiagnostic(
                    code="unsupported_dependency_source",
                    message="dependency source is not a concrete cell reference",
                    severity="error",
                    location=edge.target.normalized,
                    raw_value=edge.source.normalized,
                )
            )
            continue
        dependencies_by_target.setdefault(edge.target.normalized, []).append(edge.source.normalized)
        if index == 1 or index % 100000 == 0 or index == len(execution_edges):
            _progress(
                progress,
                f"contract inference edge scan progress execution_edges={index}/{len(execution_edges)}",
            )
    _progress(
        progress,
        f"contract inference edge scan done dependency_targets={len(dependencies_by_target)}",
    )

    input_order: list[str] = []
    input_seen: set[str] = set()
    formula_order: list[str] = []
    formula_seen: set[str] = set()
    visiting: set[str] = set()
    visited: set[str] = set()
    circular_dependency_locations: set[str] = set()
    expanded_range_dependencies: dict[str, tuple[str, ...]] = {}
    visit_calls = 0

    def dependency_refs(cell_ref: str) -> tuple[str, ...]:
        refs: list[str] = []
        for dependency in dependencies_by_target.get(cell_ref, []):
            if isinstance(dependency, str):
                refs.append(dependency)
                continue
            expanded = expanded_range_dependencies.get(dependency.normalized)
            if expanded is None:
                expanded = _expand_range_dependency(dependency)
                expanded_range_dependencies[dependency.normalized] = expanded
            refs.extend(expanded)
        return tuple(refs)

    def visit(root_ref: str) -> None:
        nonlocal visit_calls
        stack: list[tuple[str, bool]] = [(root_ref, False)]
        while stack:
            cell_ref, dependencies_processed = stack.pop()
            if cell_ref in visited:
                continue

            if dependencies_processed:
                visiting.discard(cell_ref)
                if cell_ref not in formula_seen:
                    formula_order.append(cell_ref)
                    formula_seen.add(cell_ref)
                visited.add(cell_ref)
                continue

            visit_calls += 1
            if visit_calls == 1 or visit_calls % 10000 == 0:
                _progress(
                    progress,
                    "contract inference visit progress "
                    f"calls={visit_calls} visited={len(visited)} formulas={len(formula_order)} inputs={len(input_order)}",
                )
            if cell_ref in visiting:
                if cell_ref not in circular_dependency_locations:
                    circular_dependency_locations.add(cell_ref)
                    diagnostics.append(
                        GenerationDiagnostic(
                            code="static_circular_dependency",
                            message=(
                                "selected output dependency walk encountered a static cycle; "
                                "generated runtime evaluation will fail only if the active execution path re-enters this cell"
                            ),
                            severity="warning",
                            location=cell_ref,
                        )
                    )
                continue

            cell = cell_by_ref.get(cell_ref)
            if cell is None:
                if cell_ref not in input_seen:
                    input_order.append(cell_ref)
                    input_seen.add(cell_ref)
                visited.add(cell_ref)
                continue

            if cell_ref in explicit_inputs or cell.formula is None:
                if cell_ref not in input_seen:
                    input_order.append(cell_ref)
                    input_seen.add(cell_ref)
                visited.add(cell_ref)
                continue

            visiting.add(cell_ref)
            diagnostics.extend(edge_diagnostics_by_target.get(cell_ref, ()))
            stack.append((cell_ref, True))
            for dependency_ref in reversed(dependency_refs(cell_ref)):
                if dependency_ref not in visited:
                    stack.append((dependency_ref, False))

    for output_ref in selected_outputs:
        visit(output_ref)

    _progress(
        progress,
        f"contract inference dependency walk done visited={len(visited)} formulas={len(formula_order)} inputs={len(input_order)}",
    )
    selected_expressions: dict[str, FormulaExpression] = {}
    for index, cell_ref in enumerate(formula_order, start=1):
        expression = expressions.get(cell_ref)
        if expression is None:
            diagnostics.append(
                GenerationDiagnostic(
                    code="missing_formula_expression",
                    message="inferred generated symbol has no translated formula expression",
                    severity="error",
                    location=cell_ref,
                )
            )
            continue
        selected_expressions[cell_ref] = expression
        if index == 1 or index % 10000 == 0 or index == len(formula_order):
            _progress(
                progress,
                f"contract inference expression selection progress formulas={index}/{len(formula_order)}",
            )

    constants = {cell_ref: cell_by_ref[cell_ref].raw_value if cell_ref in cell_by_ref else None for cell_ref in input_order}
    output_set = set(selected_outputs)
    symbols = tuple(
        GeneratedSymbol(cell_ref=cell_ref, symbol_name=symbol_name_for_cell_ref(cell_ref), kind="input")
        for cell_ref in input_order
    ) + tuple(
        GeneratedSymbol(
            cell_ref=cell_ref,
            symbol_name=symbol_name_for_cell_ref(cell_ref),
            kind="output" if cell_ref in output_set else "intermediate",
            raw_formula=cell_by_ref[cell_ref].formula.raw_formula if cell_by_ref[cell_ref].formula else None,
        )
        for cell_ref in formula_order
    )
    contract = GeneratedModuleContract(
        workbook_id=workbook.workbook_id,
        module_name=module_name,
        input_refs=tuple(input_order),
        output_refs=selected_outputs,
        symbols=symbols,
        include_provenance_comments=(
            inline_provenance_comment_limit is None or len(formula_order) <= inline_provenance_comment_limit
        ),
        formula_storage="lambdas"
        if inline_formula_lambda_limit is None or len(formula_order) <= inline_formula_lambda_limit
        else "expression_source",
    )
    return GeneratedContractInferenceResult(
        contract=contract,
        expressions=selected_expressions,
        constants=constants,
        diagnostics=tuple(diagnostics),
    )


def generate_python_module(
    *,
    contract: GeneratedModuleContract,
    expressions: Mapping[str, FormulaExpression],
    constants: Mapping[str, JsonValue] | None = None,
    output_path: str | Path | None = None,
    progress: Callable[[str], None] | None = None,
    runtime_progress_interval: int | None = None,
) -> GenerationResult:
    """Generate standalone Python source from translated formula expressions."""

    constants = constants or {}
    _progress(progress, f"python generation diagnostics start symbols={len(contract.symbols)}")
    diagnostics = _generation_diagnostics(contract, expressions)
    if any(diagnostic.severity == "error" for diagnostic in diagnostics):
        _progress(progress, f"python generation blocked diagnostics={len(diagnostics)}")
        return GenerationResult(contract=contract, diagnostics=tuple(diagnostics))

    _progress(progress, "python generation render start")
    source_code = _render_module(
        contract=contract,
        expressions=expressions,
        constants=constants,
        progress=progress,
        runtime_progress_interval=runtime_progress_interval,
    )
    if output_path is not None:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        _progress(progress, f"python generation write start path={path}")
        path.write_text(source_code, encoding="utf-8")
        _progress(progress, f"python generation write done bytes={path.stat().st_size}")

    _progress(progress, f"python generation done source_chars={len(source_code)} diagnostics={len(diagnostics)}")
    return GenerationResult(contract=contract, source_code=source_code, diagnostics=tuple(diagnostics))


def _generation_diagnostics(
    contract: GeneratedModuleContract,
    expressions: Mapping[str, FormulaExpression],
) -> list[GenerationDiagnostic]:
    diagnostics: list[GenerationDiagnostic] = []
    for symbol in contract.symbols:
        if symbol.kind == "input":
            continue

        expression = expressions.get(symbol.cell_ref)
        if expression is None:
            diagnostics.append(
                GenerationDiagnostic(
                    code="missing_formula_expression",
                    message="generated symbol has no translated formula expression",
                    severity="error",
                    location=symbol.cell_ref,
                    raw_value=symbol.raw_formula,
                )
            )
            continue

        if not expression.translated:
            diagnostics.append(
                GenerationDiagnostic(
                    code="unsupported_formula",
                    message="formula expression could not be generated",
                    severity="error",
                    location=symbol.cell_ref,
                    raw_value=expression.raw_formula,
                )
            )
            continue

        try:
            _render_expression(expression.root)
        except ValueError as error:
            diagnostics.append(
                GenerationDiagnostic(
                    code="formula_render_failed",
                    message=str(error),
                    severity="error",
                    location=symbol.cell_ref,
                    raw_value=expression.raw_formula,
                )
            )
    return diagnostics


def _render_module(
    *,
    contract: GeneratedModuleContract,
    expressions: Mapping[str, FormulaExpression],
    constants: Mapping[str, JsonValue],
    progress: Callable[[str], None] | None = None,
    runtime_progress_interval: int | None = None,
) -> str:
    lines = [
        '"""Generated Modelwright model.',
        "",
        f"Source workbook: {contract.workbook_id}",
        '"""',
        "",
        "import fnmatch",
        "from functools import lru_cache",
        "",
        "",
        "class _SfRangeView:",
        "    def __init__(self, sheet, min_col, min_row, max_col, max_row, get_value):",
        "        self.sheet = sheet",
        "        self.min_col = min_col",
        "        self.min_row = min_row",
        "        self.max_col = max_col",
        "        self.max_row = max_row",
        "        self._get_value = get_value",
        "        self._values = None",
        "        self._lazy_values = None",
        "        self._value_calls = 0",
        "        self._lazy_value_calls = 0",
        "",
        "    def _refs(self):",
        "        for row in range(self.min_row, self.max_row + 1):",
        "            for column in range(self.min_col, self.max_col + 1):",
        "                yield f'{self.sheet}!{_sf_column_name(column)}{row}'",
        "",
        "    def values(self):",
        "        if self._values is not None:",
        "            return self._values",
        "        values = tuple(self._get_value(ref) for ref in self._refs())",
        "        self._value_calls += 1",
        "        if self._value_calls > 1:",
        "            self._values = values",
        "        return values",
        "",
        "    def lazy_values(self):",
        "        if self._lazy_values is not None:",
        "            return self._lazy_values",
        "        values = tuple(lambda ref=ref: self._get_value(ref) for ref in self._refs())",
        "        self._lazy_value_calls += 1",
        "        if self._lazy_value_calls > 1:",
        "            self._lazy_values = values",
        "        return values",
        "",
        "",
        "def _sf_column_name(index):",
        "    name = ''",
        "    while index:",
        "        index, remainder = divmod(index - 1, 26)",
        "        name = chr(65 + remainder) + name",
        "    return name",
        "",
        "",
        "def _sf_flatten(values):",
        "    for value in values:",
        "        if isinstance(value, _SfRangeView):",
        "            yield from value.values()",
        "            continue",
        "        if isinstance(value, (list, tuple)):",
        "            yield from _sf_flatten(value)",
        "        else:",
        "            yield _sf_value(value)",
        "",
        "",
        "def _sf_flatten_lazy(values):",
        "    for value in values:",
        "        if isinstance(value, _SfRangeView):",
        "            yield from value.lazy_values()",
        "            continue",
        "        if isinstance(value, (list, tuple)):",
        "            yield from _sf_flatten_lazy(value)",
        "        else:",
        "            yield value",
        "",
        "",
        "def _sf_value(value):",
        "    return value() if callable(value) else value",
        "",
        "",
        "def _sf_number(value):",
        "    value = _sf_value(value)",
        "    return 0 if value is None else value",
        "",
        "",
        "def _sf_direct_reference(value):",
        "    value = _sf_value(value)",
        "    return 0 if value is None else value",
        "",
        "",
        "def _sf_sum_value(value):",
        "    value = _sf_value(value)",
        "    if value is None or isinstance(value, str):",
        "        return 0",
        "    return value",
        "",
        "",
        "@lru_cache(maxsize=4096)",
        "def _sf_numeric_value(value):",
        "    if isinstance(value, bool):",
        "        return None",
        "    if isinstance(value, (int, float)):",
        "        return float(value)",
        "    if isinstance(value, str):",
        "        try:",
        "            return float(value)",
        "        except ValueError:",
        "            return None",
        "    return None",
        "",
        "",
        "def _sf_average(values):",
        "    values = list(values)",
        "    return sum(values) / len(values)",
        "",
        "",
        "def _sf_iferror(value_fn, fallback):",
        "    try:",
        "        value = value_fn()",
        "    except Exception:",
        "        return fallback",
        "    if isinstance(value, str) and value.startswith('#'):",
        "        return fallback",
        "    return value",
        "",
        "",
        "def _sf_ifna(value_fn, fallback):",
        "    try:",
        "        value = value_fn()",
        "    except LookupError:",
        "        return fallback",
        "    if value == '#N/A':",
        "        return fallback",
        "    return value",
        "",
        "",
        "def _sf_coerce_criteria(raw, sample):",
        "    if isinstance(raw, str):",
        "        upper = raw.upper()",
        "        if upper == 'TRUE':",
        "            return True",
        "        if upper == 'FALSE':",
        "            return False",
        "        if isinstance(sample, str):",
        "            return raw",
        "        try:",
        "            number = float(raw)",
        "        except ValueError:",
        "            return raw",
        "        if number.is_integer():",
        "            return int(number)",
        "        return number",
        "    return raw",
        "",
        "",
        "def _sf_criteria_equal(left, right):",
        "    left_number = _sf_numeric_value(left)",
        "    right_number = _sf_numeric_value(right)",
        "    if left_number is not None and right_number is not None:",
        "        return left_number == right_number",
        "    if isinstance(left, str) and isinstance(right, str):",
        "        return left.upper() == right.upper()",
        "    return left == right",
        "",
        "",
        "def _sf_compare_criteria(value, operator, expected):",
        "    if operator == '=':",
        "        return _sf_criteria_equal(value, expected)",
        "    if operator == '<>':",
        "        return not _sf_criteria_equal(value, expected)",
        "    if operator == '>':",
        "        return value > expected",
        "    if operator == '>=':",
        "        return value >= expected",
        "    if operator == '<':",
        "        return value < expected",
        "    if operator == '<=':",
        "        return value <= expected",
        "    raise ValueError(f'unsupported criteria operator: {operator}')",
        "",
        "",
        "def _sf_criteria_matcher(criteria):",
        "    if isinstance(criteria, str):",
        "        for operator in ('>=', '<=', '<>', '>', '<', '='):",
        "            if criteria.startswith(operator):",
        "                raw_expected = criteria[len(operator):]",
        "                return lambda value: _sf_compare_criteria(value, operator, _sf_coerce_criteria(raw_expected, value))",
        "        if '*' in criteria or '?' in criteria:",
        "            return lambda value: fnmatch.fnmatchcase(str(value), criteria)",
        "        return lambda value: _sf_compare_criteria(value, '=', _sf_coerce_criteria(criteria, value))",
        "    return lambda value: _sf_compare_criteria(value, '=', criteria)",
        "",
        "",
        "def _sf_matches_criteria(value, criteria):",
        "    return _sf_criteria_matcher(criteria)(value)",
        "",
        "",
        "def _sf_lookup_equal(left, right):",
        "    if isinstance(left, str) and isinstance(right, str):",
        "        return left.upper() == right.upper()",
        "    return left == right",
        "",
        "",
        "def _sf_sumif(criteria_range, criteria, sum_range=None):",
        "    criteria_values = tuple(_sf_flatten((criteria_range,)))",
        "    sum_values = criteria_values if sum_range is None else tuple(_sf_flatten_lazy((sum_range,)))",
        "    matcher = _sf_criteria_matcher(criteria)",
        "    total = 0",
        "    for criteria_value, sum_value in zip(criteria_values, sum_values):",
        "        if matcher(criteria_value):",
        "            total += _sf_sum_value(sum_value)",
        "    return total",
        "",
        "",
        "def _sf_countif(criteria_range, criteria):",
        "    matcher = _sf_criteria_matcher(criteria)",
        "    return sum(1 for value in _sf_flatten((criteria_range,)) if matcher(value))",
        "",
        "",
        "def _sf_sumifs(sum_range, *criteria_pairs):",
        "    sum_values = tuple(_sf_flatten_lazy((sum_range,)))",
        "    criteria_ranges = [tuple(_sf_flatten((criteria_range,))) for criteria_range, _criteria in criteria_pairs]",
        "    criteria_matchers = tuple(_sf_criteria_matcher(criteria) for _range, criteria in criteria_pairs)",
        "    total = 0",
        "    for index, sum_value in enumerate(sum_values):",
        "        if all(matcher(criteria_range[index]) for criteria_range, matcher in zip(criteria_ranges, criteria_matchers)):",
        "            total += _sf_sum_value(sum_value)",
        "    return total",
        "",
        "",
        "def _sf_countifs(*criteria_pairs):",
        "    criteria_ranges = [tuple(_sf_flatten((criteria_range,))) for criteria_range, _criteria in criteria_pairs]",
        "    if not criteria_ranges:",
        "        return 0",
        "    criteria_matchers = tuple(_sf_criteria_matcher(criteria) for _range, criteria in criteria_pairs)",
        "    return sum(",
        "        1",
        "        for index in range(len(criteria_ranges[0]))",
        "        if all(matcher(criteria_range[index]) for criteria_range, matcher in zip(criteria_ranges, criteria_matchers))",
        "    )",
        "",
        "",
        "def _sf_range_lookup_enabled(range_lookup):",
        "    if isinstance(range_lookup, str):",
        "        return range_lookup.upper() not in {'FALSE', '0'}",
        "    return bool(range_lookup)",
        "",
        "",
        "def _sf_vlookup(lookup_value, table_array, col_index_num, range_lookup=True):",
        "    column_index = int(col_index_num) - 1",
        "    if column_index < 0:",
        "        raise ValueError('VLOOKUP column index must be one-based')",
        "    rows = tuple(tuple(row) for row in table_array)",
        "    if not rows:",
        "        raise LookupError('VLOOKUP table is empty')",
        "    if any(column_index >= len(row) for row in rows):",
        "        raise IndexError('VLOOKUP column index is outside the table')",
        "    if not _sf_range_lookup_enabled(range_lookup):",
        "        for row in rows:",
        "            if _sf_lookup_equal(row[0], lookup_value):",
        "                return row[column_index]",
        "        return '#N/A'",
        "    candidate = None",
        "    for row in rows:",
        "        try:",
        "            matched = row[0] <= lookup_value",
        "        except TypeError:",
        "            continue",
        "        if matched:",
        "            candidate = row",
        "        else:",
        "            break",
        "    if candidate is None:",
        "        return '#N/A'",
        "    return candidate[column_index]",
        "",
        "",
        f"def {contract.entrypoint}(inputs=None):",
        "    inputs = {} if inputs is None else dict(inputs)",
        "    _cache = {}",
        "    _range_cache = {}",
        "    _stack = []",
        "    _evaluated_count = 0",
        "    _constants = {",
    ]

    input_symbols = tuple(symbol for symbol in contract.symbols if symbol.kind == "input")
    formula_symbols = tuple(symbol for symbol in contract.symbols if symbol.kind != "input")

    for symbol in input_symbols:
        default_value = constants.get(symbol.cell_ref)
        lines.append(f"        {symbol.cell_ref!r}: {default_value!r},")

    lines.extend(
        [
            "    }",
            "",
            "    def _get(cell_ref):",
            "        nonlocal _evaluated_count",
            "        if cell_ref in inputs:",
            "            return inputs[cell_ref]",
            "        if cell_ref in _cache:",
            "            return _cache[cell_ref]",
            "        if cell_ref in _constants:",
            "            return _constants[cell_ref]",
            "        formula = _formulas.get(cell_ref)",
            "        if formula is None:",
            "            return None",
            "        if cell_ref in _stack:",
            "            cycle = _stack[_stack.index(cell_ref):] + [cell_ref]",
            "            raise RuntimeError('circular dependency during generated model execution: ' + ' -> '.join(cycle))",
            "        _stack.append(cell_ref)",
            "        try:",
            "            value = _evaluate_formula(cell_ref, formula)",
            "        finally:",
            "            _stack.pop()",
            "        _cache[cell_ref] = value",
            "        _evaluated_count += 1",
        ]
    )
    if runtime_progress_interval:
        lines.extend(
            [
                f"        if _evaluated_count == 1 or _evaluated_count % {runtime_progress_interval} == 0:",
            "            print(f'generated model progress formulas={_evaluated_count}', flush=True)",
            ]
        )
    lines.extend(
        [
            "        return value",
            "",
            "    def _range(sheet, min_col, min_row, max_col, max_row):",
            "        key = (sheet, min_col, min_row, max_col, max_row)",
            "        view = _range_cache.get(key)",
            "        if view is None:",
            "            view = _SfRangeView(sheet, min_col, min_row, max_col, max_row, _get)",
            "            _range_cache[key] = view",
            "        return view",
            "",
            "    def _table(sheet, min_col, min_row, max_col, max_row):",
            "        return tuple(",
            "            tuple(_sf_direct_reference(_get(f'{sheet}!{_sf_column_name(column)}{row}')) for column in range(min_col, max_col + 1))",
            "            for row in range(min_row, max_row + 1)",
            "        )",
            "",
        ]
    )
    if contract.formula_storage == "lambdas":
        lines.extend(
            [
                "    def _evaluate_formula(_cell_ref, formula):",
                "        return formula()",
                "",
            ]
        )
    elif contract.formula_storage == "expression_source":
        lines.extend(
            [
                "    _formula_globals = dict(globals())",
                "    _formula_globals.update({",
                "        '_get': _get,",
                "        '_range': _range,",
                "        '_table': _table,",
                "    })",
                "",
                "    def _evaluate_formula(cell_ref, formula):",
                "        code = compile(formula, f'<modelwright formula {cell_ref}>', 'eval')",
                "        return eval(code, _formula_globals)",
                "",
            ]
        )
    else:
        raise ValueError(f"unsupported formula storage: {contract.formula_storage}")
    lines.extend(
        [
            "    _formulas = {",
        ]
    )

    for index, symbol in enumerate(formula_symbols, start=1):
        if contract.include_provenance_comments:
            lines.append(f"        # {symbol.cell_ref}" + (f": {symbol.raw_formula}" if symbol.raw_formula else ""))

        expression = expressions[symbol.cell_ref]
        rendered_formula = _render_formula_root(expression.root)
        if contract.formula_storage == "lambdas":
            lines.append(f"        {symbol.cell_ref!r}: lambda: {rendered_formula},")
        else:
            lines.append(f"        {symbol.cell_ref!r}: {rendered_formula!r},")

        if index == 1 or index % 10000 == 0 or index == len(formula_symbols):
            _progress(
                progress,
                f"python generation render progress formulas={index}/{len(formula_symbols)} lines={len(lines)}",
            )
    lines.extend(
        [
            "    }",
            "    _output_refs = (",
        ]
    )
    for output_ref in contract.output_refs:
        lines.append(f"        {output_ref!r},")
    lines.extend(
        [
            "    )",
            "    return {cell_ref: _get(cell_ref) for cell_ref in _output_refs}",
        ]
    )
    lines.append("")
    return "\n".join(lines)


def _progress(progress: Callable[[str], None] | None, message: str) -> None:
    if progress is not None:
        progress(message)


def _render_formula_root(node: FormulaExpressionNode | None) -> str:
    if node is None:
        raise ValueError("cannot render missing formula expression root")
    return _render_expression(node)


def _render_expression(node: FormulaExpressionNode | None) -> str:
    if node is None:
        raise ValueError("cannot render missing formula expression root")

    if node.kind == "literal":
        return repr(node.value)
    if node.kind == "reference":
        if node.reference is None:
            raise ValueError("cannot render reference expression without reference")
        if node.reference.kind == "range":
            return _render_range_reference(node.reference)
        return f"_sf_direct_reference(_get({node.reference.normalized!r}))"
    if node.kind == "unary":
        (operand,) = node.operands
        if node.operator == "-":
            return f"(-_sf_number({_render_expression(operand)}))"
        raise ValueError(f"unsupported unary operator: {node.operator}")
    if node.kind == "binary":
        left, right = node.operands
        if node.operator == "^":
            return f"(_sf_number({_render_expression(left)}) ** _sf_number({_render_expression(right)}))"
        if node.operator == "&":
            return f"(str({_render_expression(left)}) + str({_render_expression(right)}))"
        return f"(_sf_number({_render_expression(left)}) {node.operator} _sf_number({_render_expression(right)}))"
    if node.kind == "comparison":
        left, right = node.operands
        if node.operator in {"=", "<>"}:
            return f"_sf_compare_criteria({_render_expression(left)}, {node.operator!r}, {_render_expression(right)})"
        operator = _python_comparison_operator(node.operator)
        return f"({_render_expression(left)} {operator} {_render_expression(right)})"
    if node.kind == "function_call":
        return _render_function_call(node)

    raise ValueError(f"unsupported expression kind: {node.kind}")


def _render_function_call(node: FormulaExpressionNode) -> str:
    if node.function_name == "ROUND":
        if len(node.operands) != 2:
            raise ValueError("ROUND requires two operands")
        return f"round({_render_expression(node.operands[0])}, {_render_expression(node.operands[1])})"
    if node.function_name == "IF":
        if len(node.operands) not in {2, 3}:
            raise ValueError("IF requires two or three operands")
        condition, true_value, *false_value_operands = node.operands
        false_value = false_value_operands[0] if false_value_operands else FormulaExpressionNode.literal(False)
        return f"({_render_expression(true_value)} if {_render_expression(condition)} else {_render_expression(false_value)})"
    if node.function_name == "IFERROR":
        if len(node.operands) != 2:
            raise ValueError("IFERROR requires two operands")
        value, fallback = node.operands
        return f"_sf_iferror(lambda: {_render_expression(value)}, {_render_expression(fallback)})"
    if node.function_name == "IFNA":
        if len(node.operands) != 2:
            raise ValueError("IFNA requires two operands")
        value, fallback = node.operands
        return f"_sf_ifna(lambda: {_render_expression(value)}, {_render_expression(fallback)})"
    if node.function_name == "AND":
        return f"all(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "OR":
        return f"any(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "SUM":
        return f"sum(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "MIN":
        return f"min(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "MAX":
        return f"max(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "AVERAGE":
        return f"_sf_average(_sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "CONCATENATE":
        return f"''.join(str(value) for value in _sf_flatten({_render_argument_tuple(node.operands)}))"
    if node.function_name == "SUMIF":
        if len(node.operands) not in {2, 3}:
            raise ValueError("SUMIF requires two or three operands")
        return f"_sf_sumif({_render_function_arguments(node.operands)})"
    if node.function_name == "COUNTIF":
        if len(node.operands) != 2:
            raise ValueError("COUNTIF requires two operands")
        return f"_sf_countif({_render_function_arguments(node.operands)})"
    if node.function_name == "SUMIFS":
        if len(node.operands) < 3 or len(node.operands) % 2 != 1:
            raise ValueError("SUMIFS requires a sum range followed by criteria range/criteria pairs")
        return f"_sf_sumifs({_render_criteria_function_arguments(node.operands)})"
    if node.function_name == "COUNTIFS":
        if len(node.operands) < 2 or len(node.operands) % 2 != 0:
            raise ValueError("COUNTIFS requires criteria range/criteria pairs")
        return f"_sf_countifs({_render_criteria_function_arguments(node.operands)})"
    if node.function_name == "VLOOKUP":
        if len(node.operands) not in {3, 4}:
            raise ValueError("VLOOKUP requires three or four operands")
        lookup_value, table_array, col_index_num, *range_lookup = node.operands
        rendered_arguments = [
            _render_expression(lookup_value),
            _render_table_array(table_array),
            _render_expression(col_index_num),
        ]
        if range_lookup:
            rendered_arguments.append(_render_expression(range_lookup[0]))
        return f"_sf_vlookup({', '.join(rendered_arguments)})"
    raise ValueError(f"unsupported function call: {node.function_name}")


def _render_function_arguments(operands: tuple[FormulaExpressionNode, ...]) -> str:
    return ", ".join(_render_expression(operand) for operand in operands)


def _render_criteria_function_arguments(operands: tuple[FormulaExpressionNode, ...]) -> str:
    if len(operands) < 2:
        return ""

    rendered: list[str] = []
    if len(operands) % 2 == 1:
        rendered.append(_render_expression(operands[0]))
        pair_operands = operands[1:]
    else:
        pair_operands = operands

    rendered.extend(
        f"({_render_expression(pair_operands[index])}, {_render_expression(pair_operands[index + 1])})"
        for index in range(0, len(pair_operands), 2)
    )
    return ", ".join(rendered)


def _render_argument_tuple(operands: tuple[FormulaExpressionNode, ...]) -> str:
    rendered = ", ".join(_render_expression(operand) for operand in operands)
    if len(operands) == 1:
        rendered = f"{rendered},"
    return f"({rendered})"


def _render_range_reference(reference) -> str:
    if reference.sheet is None or reference.start_cell is None or reference.end_cell is None:
        raise ValueError(f"cannot render incomplete range reference: {reference.normalized}")

    min_col, min_row, max_col, max_row = range_boundaries(f"{reference.start_cell}:{reference.end_cell}")
    return f"_range({reference.sheet!r}, {min_col}, {min_row}, {max_col}, {max_row})"


def _render_table_array(node: FormulaExpressionNode) -> str:
    if node.kind != "reference" or node.reference is None or node.reference.kind != "range":
        raise ValueError("VLOOKUP table array must be a concrete range reference")
    reference = node.reference
    if reference.sheet is None or reference.start_cell is None or reference.end_cell is None:
        raise ValueError(f"cannot render incomplete VLOOKUP table reference: {reference.normalized}")

    min_col, min_row, max_col, max_row = range_boundaries(f"{reference.start_cell}:{reference.end_cell}")
    return f"_table({reference.sheet!r}, {min_col}, {min_row}, {max_col}, {max_row})"


def _python_comparison_operator(operator: str | None) -> str:
    if operator == "=":
        return "=="
    if operator == "<>":
        return "!="
    if operator is None:
        raise ValueError("missing comparison operator")
    return operator


def _expand_range_dependency(reference) -> tuple[str, ...]:
    if reference.sheet is None or reference.start_cell is None or reference.end_cell is None:
        return ()

    min_col, min_row, max_col, max_row = range_boundaries(f"{reference.start_cell}:{reference.end_cell}")
    return tuple(
        f"{reference.sheet}!{get_column_letter(column)}{row}"
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    )
